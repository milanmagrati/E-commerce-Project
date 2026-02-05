from urllib import request
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.views.decorators.http import require_POST, require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Prefetch
from django.http import JsonResponse, HttpResponse, Http404
from datetime import datetime, timedelta

import requests
from .models import (Product, Order, OrderItem, Category, Customer, 
                     ProductVariation, ProductImage, ProductVariantOption,
                     OrderActivityLog, StockIn, City, StockInItem)
from decimal import Decimal, InvalidOperation
import json
from .forms import ProductForm, ProductVariationForm, ProductVariationFormSet, CustomerForm, OrderForm
from django.db import IntegrityError, transaction, connection
from django.utils import timezone
import traceback
import uuid, os
from django.core.files.storage import default_storage
from django.core.files.base import File
from django.conf import settings
from django.utils.text import slugify
from .models import ReturnRequest, ReturnItem, ReturnActivityLog, Dispatch, DispatchItem

# ✅ IMPORT DECORATORS
from accounts.decorators import permission_required, admin_only

# ✅ GET CUSTOM USER MODEL
User = get_user_model()


def fix_order_decimals(order):
    """Fix any NULL decimal values in order"""
    if order.discount_amount is None:
        order.discount_amount = Decimal("0")
    if order.shipping_charge is None:
        order.shipping_charge = Decimal("0")
    if order.tax_percent is None:
        order.tax_percent = Decimal("0")
    if order.total_amount is None:
        order.total_amount = Decimal("0")
    
    # ✅ ENHANCED PARTIAL PAYMENT DECIMAL FIXES
    if order.partial_amount_paid is None:
        order.partial_amount_paid = Decimal("0")
    if order.remaining_amount is None:
        # Auto-calculate remaining if partial payment
        if order.is_partial_payment or order.payment_status == 'partial':
            order.remaining_amount = order.total_amount - order.partial_amount_paid
            if order.remaining_amount < Decimal("0"):
                order.remaining_amount = Decimal("0")
        else:
            order.remaining_amount = Decimal("0")
    
    # ✅ ENSURE is_partial_payment IS SYNCED WITH PAYMENT_STATUS
    if order.payment_status == 'partial' and not order.is_partial_payment:
        order.is_partial_payment = True
    elif order.payment_status != 'partial' and order.is_partial_payment:
        order.is_partial_payment = False
    
    return order



def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def dashboard_view(request):
    # System-wide product and order data (show counts to staff like warehouse)
    products = Product.objects.filter(is_deleted=False)
    orders = Order.objects.all()
    
    # Statistics
    total_products = products.count()
    total_orders = orders.count()
    pending_orders = orders.filter(order_status='pending').count()
    processing_orders = orders.filter(order_status='processing').count()
    shipped_orders = orders.filter(order_status='shipped').count()
    delivered_orders = orders.filter(order_status='delivered').count()
    
    total_revenue = orders.filter(payment_status='paid').aggregate(
        total=Sum('total_amount'))['total'] or 0
    
    # Recent orders
    recent_orders = orders.order_by('-created_at')[:5]
    
    # Low stock products
    low_stock_products = products.filter(stock__lte=10, stock__gt=0).order_by('stock')[:5]
    
    # Monthly sales data for chart (last 6 months)
    monthly_sales = []
    for i in range(5, -1, -1):
        date = timezone.now() - timedelta(days=30*i)
        month_name = date.strftime('%b %Y')
        month_start = date.replace(day=1)
        
        if i > 0:
            next_month = (date.replace(day=28) + timedelta(days=4)).replace(day=1)
        else:
            next_month = timezone.now() + timedelta(days=1)
        
        sales = orders.filter(
            created_at__gte=month_start,
            created_at__lt=next_month,
            payment_status='paid'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        monthly_sales.append({
            'month': month_name,
            'sales': float(sales)
        })
    
    context = {
        'total_products': total_products,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        'total_revenue': total_revenue,
        'recent_orders': recent_orders,
        'low_stock_products': low_stock_products,
        'monthly_sales': json.dumps(monthly_sales),
    }
    
    return render(request, 'dashboard.html', context)
@login_required
@permission_required('can_view_products')
def products_view(request):
    # Check for a flag set by product_add to clear any client-side product drafts
    clear_product_draft = request.session.pop('clear_product_draft', False)
    # attach to request for template access
    request.clear_product_draft = clear_product_draft
    """Products list with search, filters, and date range"""
    products = Product.objects.filter(
        is_deleted=False
    ).select_related('category').order_by('-created_at')    
    # Search functionality
    search_query = request.GET.get("search", "")
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(slug__icontains=search_query)
        )
    
    # Category filter
    category_filter = request.GET.get("category", "")
    if category_filter:
        products = products.filter(category__slug=category_filter)
    
    # Status filter
    status_filter = request.GET.get("status", "")
    if status_filter == "active":
        products = products.filter(is_active=True)
    elif status_filter == "inactive":
        products = products.filter(is_active=False)
    
    # Stock filter
    stock_filter = request.GET.get("stock", "")
    if stock_filter == "in_stock":
        products = products.filter(stock__gt=10)
    elif stock_filter == "low_stock":
        products = products.filter(stock__lte=10, stock__gt=0)
    elif stock_filter == "out_of_stock":
        products = products.filter(stock=0)
    
    # Date Range Filter
    date_filter = request.GET.get("date_range", "")
    today = timezone.now().date()
    
    if date_filter == "today":
        products = products.filter(created_at__date=today)
    elif date_filter == "yesterday":
        yesterday = today - timedelta(days=1)
        products = products.filter(created_at__date=yesterday)
    elif date_filter == "last_7_days":
        start_date = today - timedelta(days=7)
        products = products.filter(created_at__date__gte=start_date)
    elif date_filter == "last_30_days":
        start_date = today - timedelta(days=30)
        products = products.filter(created_at__date__gte=start_date)
    elif date_filter == "this_month":
        products = products.filter(
            created_at__year=today.year,
            created_at__month=today.month
        )
    elif date_filter == "last_month":
        first_day_this_month = today.replace(day=1)
        last_month = first_day_this_month - timedelta(days=1)
        products = products.filter(
            created_at__year=last_month.year,
            created_at__month=last_month.month
        )
    elif date_filter == "this_year":
        products = products.filter(created_at__year=today.year)
    elif date_filter == "custom":
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
                products = products.filter(created_at__date__gte=start_date_obj)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
                products = products.filter(created_at__date__lte=end_date_obj)
            except ValueError:
                pass
    
    # Calculate statistics (on all user products, not filtered)
    all_products = Product.objects.filter(is_deleted=False)
    active_count = all_products.filter(is_active=True).count()
    low_stock_count = all_products.filter(stock__lte=10, stock__gt=0).count()
    out_of_stock_count = all_products.filter(stock=0).count()
    
    categories = Category.objects.all()
    
    context = {
        "products": products,
        "categories": categories,
        "search_query": search_query,
        "category_filter": category_filter,
        "status_filter": status_filter,
        "stock_filter": stock_filter,
        "date_filter": date_filter,
        "start_date": request.GET.get("start_date", ""),
        "end_date": request.GET.get("end_date", ""),
        "active_count": active_count,
        "low_stock_count": low_stock_count,
        "out_of_stock_count": out_of_stock_count,
        "clear_product_draft": getattr(request, 'clear_product_draft', False),
    }
    
    return render(request, "products.html", context)


@login_required
@permission_required('can_delete_products')
def products_bulk_action(request):
    """Handle bulk actions on products"""
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        action = request.POST.get('bulk_action')
        
        if not product_ids:
            messages.error(request, 'No products selected!')
            return redirect('products')
        
        try:
            # ✅ UPDATED: Remove user filter - show all products
            products = Product.objects.filter(
                id__in=product_ids,
                is_deleted=False
            )
            count = products.count()
            
            if count == 0:
                messages.error(request, 'No valid products found!')
                return redirect('products')
            
            if action == 'delete':
                products.update(is_deleted=True, deleted_at=timezone.now())
                messages.success(request, f'{count} product(s) moved to trash!')
                
            elif action == 'activate':
                products.update(is_active=True)
                messages.success(request, f'{count} product(s) activated!')
                
            elif action == 'deactivate':
                products.update(is_active=False)
                messages.success(request, f'{count} product(s) deactivated!')
                
            elif action == 'increase_price':
                percentage = request.POST.get('percentage')
                if percentage:
                    try:
                        percentage = Decimal(percentage) / 100
                        for product in products:
                            product.price = product.price * (1 + percentage)
                            product.save()
                        messages.success(request, f'Price increased by {float(percentage)*100}% for {count} product(s)!')
                    except (ValueError, TypeError):
                        messages.error(request, 'Invalid percentage value!')
                else:
                    messages.error(request, 'Please provide a percentage!')
                    
            elif action == 'decrease_price':
                percentage = request.POST.get('percentage')
                if percentage:
                    try:
                        percentage = Decimal(percentage) / 100
                        for product in products:
                            new_price = product.price * (1 - percentage)
                            if new_price > 0:
                                product.price = new_price
                                product.save()
                        messages.success(request, f'Price decreased by {float(percentage)*100}% for {count} product(s)!')
                    except (ValueError, TypeError):
                        messages.error(request, 'Invalid percentage value!')
                else:
                    messages.error(request, 'Please provide a percentage!')
                    
            else:
                messages.error(request, 'Invalid action selected!')
                
        except Exception as e:
            messages.error(request, f'Error performing bulk action: {str(e)}')
    
    return redirect('products')

@login_required
@permission_required('can_create_products')

def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        
        if form.is_valid():
            product = form.save(commit=False)
            product.user = request.user
            product.save()

            # If a temporary uploaded image exists (from a previous failed validation), attach it to the saved product
            try:
                temp_path = request.session.pop('temp_product_image', None)
                if temp_path and default_storage.exists(temp_path):
                    with default_storage.open(temp_path, 'rb') as tf:
                        product.image.save(os.path.basename(temp_path), File(tf), save=True)
                    try:
                        default_storage.delete(temp_path)
                    except Exception:
                        pass
            except Exception:
                pass
            
            # Save variant options if variable product
            if product.product_type == 'variable':
                variant_options = form.cleaned_data.get('variant_options')
                size_options = form.cleaned_data.get('size_options')
                
                if variant_options:
                    ProductVariantOption.objects.create(
                        product=product,
                        option_name='Variant',
                        option_values=variant_options
                    )
                
                if size_options:
                    ProductVariantOption.objects.create(
                        product=product,
                        option_name='Size',
                        option_values=size_options
                    )
                
                # Handle variation formset
                formset = ProductVariationFormSet(request.POST, request.FILES, instance=product)
                if formset.is_valid():
                    formset.save()
                else:
                    # Preserve uploaded image if any by moving to a temp location before deleting the product
                    temp_path = None
                    temp_url = None
                    try:
                        if product.image:
                            # Copy the saved product image to temp_uploads
                            base = os.path.basename(product.image.name)
                            temp_name = f"temp_uploads/{uuid.uuid4().hex}_{base}"
                            with product.image.open('rb') as f:
                                temp_path = default_storage.save(temp_name, f)
                                temp_url = default_storage.url(temp_path)
                                # store in session so it can be reused across requests
                                request.session['temp_product_image'] = temp_path
                    except Exception as e:
                        print('Temp image save failed:', e)

                    product.delete()
                    messages.error(request, 'Error in product variations. Please check the form.')
                    ctx = {
                        'form': form,
                        'formset': formset,
                        'action': 'Add',
                        'current_step': 2
                    }
                    if temp_url:
                        ctx['temp_image_url'] = temp_url
                        ctx['temp_image_path'] = temp_path

                    return render(request, 'dashboard/product_form.html', ctx)
            
            # Handle gallery images
            gallery_images = request.FILES.getlist('gallery_images')
            for img in gallery_images:
                ProductImage.objects.create(product=product, image=img)
            
            messages.success(request, f'Product "{product.name}" created successfully!')
            # Clean up any temporary uploaded image saved in session
            temp_to_remove = request.session.pop('temp_product_image', None)
            if temp_to_remove and default_storage.exists(temp_to_remove):
                try:
                    default_storage.delete(temp_to_remove)
                except Exception:
                    pass
            # Mark session to clear any product draft stored in localStorage on the products page
            request.session['clear_product_draft'] = True
            return redirect('products')
        else:
            formset = ProductVariationFormSet(request.POST, request.FILES)

            # If user uploaded an image but form validation failed, persist it to temp storage
            try:
                if 'image' in request.FILES:
                    uploaded = request.FILES['image']
                    temp_name = f"temp_uploads/{uuid.uuid4().hex}_{uploaded.name}"
                    temp_path = default_storage.save(temp_name, uploaded)
                    temp_url = default_storage.url(temp_path)
                    request.session['temp_product_image'] = temp_path
                else:
                    temp_path = request.session.get('temp_product_image')
                    temp_url = default_storage.url(temp_path) if temp_path and default_storage.exists(temp_path) else None
            except Exception as e:
                temp_path = None
                temp_url = None
                print('Temp save error:', e)

            messages.error(request, 'Please correct the errors below.')
            return render(request, 'product_form.html', {
                'form': form,
                'formset': formset,
                'action': 'Add',
                'current_step': 1,
                'temp_image_url': temp_url,
                'temp_image_path': temp_path,
            })
    else:
        form = ProductForm()
        formset = ProductVariationFormSet()
    
    # Respect cleared=1 param from Clear & Start Fresh to delete any temp uploaded image
    if 'cleared' in request.GET:
        try:
            tmp_pop = request.session.pop('temp_product_image', None)
            if tmp_pop and default_storage.exists(tmp_pop):
                default_storage.delete(tmp_pop)
        except Exception:
            pass

    # If there's a temp image in session (from previous failed upload), pass it to the template
    temp_path = request.session.get('temp_product_image')
    temp_url = None
    try:
        if temp_path and default_storage.exists(temp_path):
            temp_url = default_storage.url(temp_path)
        else:
            temp_path = None
    except Exception:
        temp_path = None

    return render(request, 'product_form.html', {
        'form': form,
        'formset': formset,
        'action': 'Add',
        'current_step': 1,
        'temp_image_url': temp_url,
        'temp_image_path': temp_path,
    })
@permission_required('can_edit_products')
def product_edit(request, product_id):
    product = get_object_or_404(Product, pk=product_id, is_deleted=False)
    
    # Get existing variant options
    variant_option = product.variant_options.filter(option_name='Variant').first()
    size_option = product.variant_options.filter(option_name='Size').first()
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        formset = ProductVariationFormSet(request.POST, request.FILES, instance=product)
        
        if form.is_valid():
            product = form.save()

            # If a temporary uploaded image exists (from previous failed validation), attach it to the product
            try:
                temp_path = request.session.pop('temp_product_image', None)
                if temp_path and default_storage.exists(temp_path) and not product.image:
                    with default_storage.open(temp_path, 'rb') as tf:
                        product.image.save(os.path.basename(temp_path), File(tf), save=True)
                    try:
                        default_storage.delete(temp_path)
                    except Exception:
                        pass
            except Exception:
                pass
            
            # Update variant options
            if product.product_type == 'variable':
                variant_options = form.cleaned_data.get('variant_options')
                size_options = form.cleaned_data.get('size_options')
                
                # Update or create variant option
                if variant_options:
                    if variant_option:
                        variant_option.option_values = variant_options
                        variant_option.save()
                    else:
                        ProductVariantOption.objects.create(
                            product=product,
                            option_name='Variant',
                            option_values=variant_options
                        )
                
                # Update or create size option
                if size_options:
                    if size_option:
                        size_option.option_values = size_options
                        size_option.save()
                    else:
                        ProductVariantOption.objects.create(
                            product=product,
                            option_name='Size',
                            option_values=size_options
                        )
                
                if formset.is_valid():
                    formset.save()
            
            # Handle gallery images
            gallery_images = request.FILES.getlist('gallery_images')
            for img in gallery_images:
                ProductImage.objects.create(product=product, image=img)
            
            messages.success(request, f'Product "{product.name}" updated successfully!')
            # Clean up any temporary uploaded image saved in session
            temp_to_remove = request.session.pop('temp_product_image', None)
            if temp_to_remove and default_storage.exists(temp_to_remove):
                try:
                    default_storage.delete(temp_to_remove)
                except Exception:
                    pass
            return redirect('products')
        else:
            messages.error(request, 'Please correct the errors below.')
            # If user uploaded an image but form validation failed, persist it to temp storage
            try:
                if 'image' in request.FILES:
                    uploaded = request.FILES['image']
                    temp_name = f"temp_uploads/{uuid.uuid4().hex}_{uploaded.name}"
                    temp_path = default_storage.save(temp_name, uploaded)
                    request.session['temp_product_image'] = temp_path
            except Exception as e:
                print('Temp save error on edit:', e)
    else:
        form = ProductForm(instance=product, initial={
            'variant_options': variant_option.option_values if variant_option else '',
            'size_options': size_option.option_values if size_option else ''
        })
        formset = ProductVariationFormSet(instance=product)
    
    # If there's a temp image in session (from previous failed upload), pass it to the template
    temp_path = request.session.get('temp_product_image')
    temp_url = None
    try:
        if temp_path and default_storage.exists(temp_path):
            temp_url = default_storage.url(temp_path)
        else:
            temp_path = None
    except Exception:
        temp_path = None

    return render(request, 'product_form.html', {
        'form': form,
        'formset': formset,
        'product': product,
        'action': 'Edit',
        'current_step': 1,
        'temp_image_url': temp_url,
        'temp_image_path': temp_path
    })


@login_required
@permission_required('can_view_products')
def product_detail(request, product_id):
    product = get_object_or_404(Product, pk=product_id, user=request.user)
    
    try:
        profit = (product.price or 0) - (product.cost_price or 0)
    except Exception:
        profit = 0

    context = {
        "product": product,
        "profit": profit,
    }
    
    # Handle variation updates
    if request.method == 'POST':
        if 'delete_variation_id' in request.POST:
            # Delete specific variation
            variation_id = request.POST.get('delete_variation_id')
            variation = get_object_or_404(ProductVariation, id=variation_id, product=product)
            variation.delete()
            messages.success(request, f'Variation "{variation.sku}" deleted successfully!')
            return redirect('product_detail', product_id=product_id)
        
        elif 'update_variations' in request.POST:
            # Update or create variations
            variations_data = {}
            
            # Parse variations from POST data
            for key, value in request.POST.items():
                if key.startswith('variations['):
                    # Extract index and field name
                    parts = key.replace('variations[', '').replace(']', '').split('[')
                    if len(parts) == 2:
                        index, field = parts
                        if index not in variations_data:
                            variations_data[index] = {}
                        variations_data[index][field] = value
            
            # Process each variation
            for index, var_data in variations_data.items():
                variation_id = var_data.get('id')
                sku = var_data.get('sku')
                price = var_data.get('price')
                stock = var_data.get('stock')
                status = var_data.get('status', 'active')
                is_active = f'variations[{index}][is_active]' in request.POST
                
                if variation_id:
                    # Update existing
                    variation = ProductVariation.objects.get(id=variation_id, product=product)
                    variation.sku = sku
                    variation.price = price
                    variation.stock = stock
                    variation.status = status
                    variation.is_active = is_active
                else:
                    # Create new
                    variation = ProductVariation.objects.create(
                        product=product,
                        sku=sku,
                        price=price,
                        stock=stock,
                        status=status,
                        is_active=is_active
                    )
                
                # Handle image upload
                image_key = f'variations[{index}][image]'
                if image_key in request.FILES:
                    variation.image = request.FILES[image_key]
                
                variation.save()
            
            messages.success(request, 'Variations updated successfully!')
            return redirect('product_detail', product_id=product_id)
    
    # Get product data
    product_images = product.images.all()
    order_items = product.orderitem_set.all()[:10]
    
    # Calculate profit margin
    profit_margin = 0
    if product.cost_price and product.cost_price > 0:
        profit_margin = ((product.price - product.cost_price) / product.price) * 100
    
    context = {
        'product': product,
        'product_images': product_images,
        'order_items': order_items,
        'profit_margin': profit_margin,
    }
    
    return render(request, 'product_detail.html', context)


@login_required
@permission_required('can_view_products')
def products_trash(request):
    """View trashed products"""
    if request.user.is_superuser or request.user.role == 'admin':
        trashed_products = Product.objects.filter(
            is_deleted=True
        ).select_related('category').order_by('-deleted_at')
    else:
        trashed_products = Product.objects.filter(
            user=request.user,
            is_deleted=True
        ).select_related('category').order_by('-deleted_at')

    # Search functionality
    search_query = request.GET.get("search", "")
    if search_query:
        trashed_products = trashed_products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(slug__icontains=search_query)
        )
    
    # Category filter
    category_filter = request.GET.get("category", "")
    if category_filter:
        trashed_products = trashed_products.filter(category__slug=category_filter)
    
    categories = Category.objects.all()
    
    context = {
        "trashed_products": trashed_products,
        "categories": categories,
        "search_query": search_query,
        "category_filter": category_filter,
    }
    
    return render(request, "products_trash.html", context)


@login_required
@permission_required('can_delete_products')
def product_move_to_trash(request, product_id):
    """Move product to trash (soft delete)"""
    product = get_object_or_404(Product, id=product_id, is_deleted=False)
    
    if request.method == 'POST':
        product_name = product.name
        product.is_deleted = True
        product.deleted_at = timezone.now()
        product.save()
        
        messages.success(request, f'Product "{product_name}" moved to trash successfully!')
        return redirect('products')
    
    return redirect('product_detail', product_id=product_id)


@login_required
@permission_required('can_delete_products')
def product_restore(request, product_id):
    """Restore product from trash"""
    product = get_object_or_404(Product, id=product_id, user=request.user, is_deleted=True)
    
    if request.method == 'POST':
        product_name = product.name
        product.is_deleted = False
        product.deleted_at = None
        product.save()
        
        messages.success(request, f'Product "{product_name}" restored successfully!')
        return redirect('products_trash')
    
    return redirect('products_trash')


@login_required
@permission_required('can_delete_products')
def product_permanent_delete(request, product_id):
    """Permanently delete product"""
    product = get_object_or_404(Product, id=product_id, user=request.user, is_deleted=True)
    
    if request.method == 'POST':
        product_name = product.name
        product.delete()
        
        messages.success(request, f'Product "{product_name}" permanently deleted!')
        return redirect('products_trash')
    
    return redirect('products_trash')


@login_required
@permission_required('can_delete_products')
def products_trash_bulk_action(request):
    """Handle bulk actions on trashed products"""
    if request.method == "POST":
        product_ids = request.POST.getlist("product_ids")
        action = request.POST.get("bulk_action")
        
        if not product_ids:
            messages.error(request, "No products selected!")
            return redirect('products_trash')
        
        try:
            products = Product.objects.filter(
                id__in=product_ids, 
                user=request.user, 
                is_deleted=True
            )
            count = products.count()
            
            if count == 0:
                messages.error(request, "No valid products found!")
                return redirect('products_trash')
            
            if action == "restore":
                products.update(is_deleted=False, deleted_at=None)
                messages.success(request, f"✅ {count} product(s) restored successfully!")
                
            elif action == "permanent_delete":
                products.delete()
                messages.success(request, f"✅ {count} product(s) permanently deleted!")
                
            else:
                messages.error(request, "Invalid action selected!")
                
        except Exception as e:
            messages.error(request, f"Error performing bulk action: {str(e)}")
            
    return redirect('products_trash')


@login_required
@permission_required('can_delete_products')
def empty_trash(request):
    """Empty all trashed products"""
    if request.method == 'POST':
        trashed_products = Product.objects.filter(user=request.user, is_deleted=True)
        count = trashed_products.count()
        
        if count > 0:
            trashed_products.delete()
            messages.success(request, f'✅ Trash emptied! {count} product(s) permanently deleted.')
        else:
            messages.info(request, 'Trash is already empty.')
        
        return redirect('products_trash')
    
    return redirect('products_trash')


@login_required
@permission_required('can_edit_products')
def delete_product_image(request, image_id):
    """Delete a product gallery image"""
    image = get_object_or_404(ProductImage, id=image_id, product__user=request.user)
    product_id = image.product.id
    product_name = image.product.name
    
    # Delete the image
    image.delete()
    
    messages.success(request, f'Image deleted from "{product_name}" gallery successfully!')
    return redirect('product_detail', product_id=product_id)


@login_required
@permission_required('can_edit_products')
def set_featured_image(request, image_id):
    """Set an image as featured in the gallery"""
    image = get_object_or_404(ProductImage, id=image_id, product__user=request.user)
    
    # Unset all other featured images for this product
    ProductImage.objects.filter(product=image.product).update(is_featured=False)
    
    # Set this image as featured
    image.is_featured = True
    image.save()
    
    messages.success(request, f'Featured image updated for "{image.product.name}"!')
    return redirect('product_detail', product_id=image.product.id)


@login_required
@permission_required('can_edit_products')
def upload_product_images(request, product_id):
    """Upload multiple images to product gallery"""
    product = get_object_or_404(Product, id=product_id, user=request.user)
    
    if request.method == 'POST':
        gallery_images = request.FILES.getlist('images')
        
        if gallery_images:
            # Get current max order
            max_order = ProductImage.objects.filter(product=product).count()
            
            for idx, image in enumerate(gallery_images):
                ProductImage.objects.create(
                    product=product,
                    image=image,
                    order=max_order + idx,
                    alt_text=f"{product.name} - Gallery Image {max_order + idx + 1}"
                )
            
            messages.success(request, f'{len(gallery_images)} image(s) uploaded successfully to "{product.name}" gallery!')
        else:
            messages.warning(request, 'No images were selected.')
        
        return redirect('product_detail', product_id=product.id)
    
    return redirect('product_detail', product_id=product.id)


@login_required
@permission_required('can_edit_products')
def reorder_product_images(request, product_id):
    """Reorder product gallery images via AJAX"""
    if request.method == 'POST':
        product = get_object_or_404(Product, id=product_id, user=request.user)
        
        try:
            order_data = json.loads(request.body)
            
            for item in order_data:
                image_id = item.get('id')
                new_order = item.get('order')
                
                ProductImage.objects.filter(
                    id=image_id, 
                    product=product
                ).update(order=new_order)
            
            return JsonResponse({'success': True, 'message': 'Images reordered successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
@permission_required('can_view_orders')
def orders_view(request):
    orders = Order.objects.filter(user=request.user)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        orders = orders.filter(order_status=status_filter)
    
    context = {
        'orders': orders,
        'status_filter': status_filter,
    }
    
    return render(request, 'orders.html', context)




@login_required
def chart_data(request):
    # Get sales data for charts
    today = datetime.now().date()
    
    # Last 7 days sales
    daily_sales = []
    for i in range(7):
        date = today - timedelta(days=i)
        sales = Order.objects.filter(
            user=request.user,
            payment_status='paid',
            created_at__date=date
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        daily_sales.insert(0, {
            'date': date.strftime('%d %b'),
            'sales': float(sales)
        })
    
    return JsonResponse({
        'daily_sales': daily_sales
    })


@login_required
@permission_required('can_view_customers')
def customers_view(request):
    customers = Customer.objects.all().order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) | 
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Add order count and total spent for each customer
    customers_data = []
    for customer in customers:
        orders = Order.objects.filter(customer_email=customer.email)
        total_orders = orders.count()
        total_spent = orders.filter(payment_status='paid').aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00')
        
        customers_data.append({
            'customer': customer,
            'total_orders': total_orders,
            'total_spent': total_spent,
            'last_order': orders.first(),
        })
    
    context = {
        'customers_data': customers_data,
        'search_query': search_query,
    }
    # Ensure older view renders the consolidated customers_list template
    context['customer_type'] = ''
    return render(request, 'customers_list.html', context)


@login_required
@permission_required('can_view_customers')
def customer_detail(request, customer_id):
    """View customer details"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    # ✅ CORRECT: Query orders by customer email and phone
    customer_orders = Order.objects.filter(
        Q(customer_email=customer.email) | Q(customer_phone=customer.phone)
    ).order_by('-created_at')
    
    # Calculate statistics
    total_orders = customer_orders.count()
    total_spent = customer_orders.filter(payment_status='paid').aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')
    
    pending_orders = customer_orders.filter(order_status='pending').count()
    delivered_orders = customer_orders.filter(order_status='delivered').count()
    
    # Get last order date
    last_order = customer_orders.first()
    last_order_date = last_order.created_at if last_order else None
    
    context = {
        'customer': customer,
        'customer_orders': customer_orders,
        'total_orders': total_orders,
        'total_spent': total_spent,
        'pending_orders': pending_orders,
        'delivered_orders': delivered_orders,
        'last_order_date': last_order_date,
    }
    
    return render(request, 'customer_detail.html', context)

@login_required
@permission_required('can_delete_customers')
def customers_bulk_action(request):
    """Handle bulk actions on customers"""
    if request.method == 'POST':
        customer_ids = request.POST.getlist('customer_ids')
        action = request.POST.get('bulk_action')
        
        if not customer_ids:
            messages.error(request, 'No customers selected!')
            return redirect('customers_list')
        
        try:
            customers = Customer.objects.filter(id__in=customer_ids)
            count = customers.count()
            
            if count == 0:
                messages.error(request, 'No valid customers found!')
                return redirect('customers_list')
            
            if action == 'delete':
                customers.delete()
                messages.success(request, f'✅ {count} customer(s) deleted successfully!')
                
            elif action == 'activate':
                customers.update(is_active=True)
                messages.success(request, f'✅ {count} customer(s) activated!')
                
            elif action == 'deactivate':
                customers.update(is_active=False)
                messages.success(request, f'✅ {count} customer(s) deactivated!')
                
            elif action == 'change_type_retail':
                customers.update(customer_type='retail')
                messages.success(request, f'✅ {count} customer(s) changed to Retail!')
                
            elif action == 'change_type_wholesale':
                customers.update(customer_type='wholesale')
                messages.success(request, f'✅ {count} customer(s) changed to Wholesale!')
                
            elif action == 'change_type_vip':
                customers.update(customer_type='vip')
                messages.success(request, f'✅ {count} customer(s) changed to VIP!')
                
            elif action == 'export':
                # Export to Excel
                wb = Workbook()
                ws = wb.active
                ws.title = 'Customers'
                
                # Styling
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                
                # Headers
                headers = ['Name', 'Email', 'Phone', 'Alternate Phone', 'City', 'Address', 'Type', 'Total Orders', 'Total Spent']
                ws.append(headers)
                
                # Style header row
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Data rows
                for customer in customers:
                    orders = customer.orders.all()
                    total_spent = orders.filter(payment_status='paid').aggregate(
                        total=Sum('total_amount'))['total'] or Decimal('0.00')
                    
                    ws.append([
                        customer.name,
                        customer.email or '',
                        customer.phone,
                        customer.alternate_phone or '',
                        customer.city,
                        customer.address,
                        customer.customer_type.upper(),
                        orders.count(),
                        float(total_spent)
                    ])
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Create response
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename=customers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                wb.save(response)
                return response
                
            else:
                messages.error(request, 'Invalid action selected!')
                
        except Exception as e:
            messages.error(request, f'❌ Error performing bulk action: {str(e)}')
    
    return redirect('customers_list')


@login_required
@admin_only
def category_list(request):
    categories = Category.objects.all().order_by('name')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        slug = request.POST.get('slug')
        
        if name and slug:
            Category.objects.create(name=name, slug=slug)
            messages.success(request, f'Category "{name}" added successfully!')
            return redirect('category_list')
    
    context = {
        'categories': categories,
    }
    return render(request, 'category_list.html', context)


@login_required
@admin_only
def category_delete(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Category "{category_name}" deleted successfully!')
        return redirect('category_list')
    
    return redirect('category_list')


@login_required
@permission_required('can_edit_products')
def product_variations(request, product_id):
    product = get_object_or_404(Product, id=product_id, user=request.user)
    
    if product.product_type != 'variable':
        messages.warning(request, 'This product is not a variable product.')
        return redirect('products')
    
    variations = ProductVariation.objects.filter(product=product)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_variation':
            sku = request.POST.get('sku', '').strip()
            price = request.POST.get('price') or None
            stock = request.POST.get('stock', 0)
            
            # Validate SKU
            if not sku:
                messages.error(request, 'SKU is required!')
                return redirect('product_variations', product_id=product.id)
            
            # Check if SKU already exists
            if ProductVariation.objects.filter(sku=sku).exists():
                messages.error(request, f'SKU "{sku}" already exists! Please use a unique SKU.')
                return redirect('product_variations', product_id=product.id)
            
            try:
                variation = ProductVariation.objects.create(
                    product=product,
                    sku=sku,
                    price=price,
                    stock=stock
                )
                
                if 'variation_image' in request.FILES:
                    variation.image = request.FILES['variation_image']
                    variation.save()
                
                messages.success(request, f'Variation "{sku}" added successfully!')
                return redirect('product_variations', product_id=product.id)
                
            except IntegrityError:
                messages.error(request, f'SKU "{sku}" already exists! Please use a unique SKU.')
                return redirect('product_variations', product_id=product.id)
            except Exception as e:
                messages.error(request, f'Error creating variation: {str(e)}')
                return redirect('product_variations', product_id=product.id)
    
    context = {
        'product': product,
        'variations': variations,
    }
    return render(request, 'product_variations.html', context) 


@login_required
@permission_required('can_delete_products')
def variation_delete(request, variation_id):
    variation = get_object_or_404(ProductVariation, id=variation_id, product__user=request.user)
    product_id = variation.product.id
    variation_sku = variation.sku
    variation.delete()
    messages.success(request, f'Variation "{variation_sku}" deleted successfully!')
    return redirect('product_variations', product_id=product_id)


# Attributes feature removed: attribute management was removed as requested. Views and templates related to attributes were deleted to simplify product handling.


@login_required
@permission_required('can_view_customers')
def customers_list(request):
    """List all customers with search and filter"""
    customers = Customer.objects.all().order_by('-created_at')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Filter by type
    customer_type = request.GET.get('type', '')
    if customer_type:
        customers = customers.filter(customer_type=customer_type)
    
    # Add statistics
    customers_data = []
    for customer in customers:
      orders = customer.orders.all()
      customers_data.append({
            'customer': customer,
            'total_orders': orders.count(),
            'total_spent': orders.filter(payment_status='paid').aggregate(
                total=Sum('total_amount'))['total'] or Decimal('0.00'),
            'last_order': orders.first(),
        })
    
    context = {
        'customers_data': customers_data,
        'search_query': search_query,
        'customer_type': customer_type,
    }
    return render(request, 'customers_list.html', context)


@login_required
@permission_required('can_create_customers')
def customer_add(request):
    """Add new customer"""
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.user = request.user
            customer.save()
            messages.success(request, f'Customer "{customer.name}" added successfully!')
            return redirect('customers_list')
        else:
            # Print form errors for debugging
            print("Form Errors:", form.errors)
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomerForm()
    
    context = {'form': form, 'action': 'Add'}
    return render(request, 'customer_form.html', context)


@login_required
@permission_required('can_edit_customers')
def customer_edit(request, customer_id):
    """Edit existing customer"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f'Customer "{customer.name}" updated successfully!')
            return redirect('customer_detail', customer_id=customer.id)
    else:
        form = CustomerForm(instance=customer)
    
    context = {'form': form, 'customer': customer, 'action': 'Edit'}
    return render(request, 'customer_form.html', context)


@login_required
def customer_detail(request, customer_id):
    """View customer details"""
    customer = get_object_or_404(Customer, id=customer_id)
    orders = customer.orders.all().order_by('-created_at')
    
    # Statistics
    total_orders = orders.count()
    total_spent = orders.filter(payment_status='paid').aggregate(
        total=Sum('total_amount'))['total'] or Decimal('0.00')
    pending_orders = orders.filter(order_status='pending').count()
    
    context = {
        'customer': customer,
        'orders': orders,
        'total_orders': total_orders,
        'total_spent': total_spent,
        'pending_orders': pending_orders,
    }
    return render(request, 'customer_detail.html', context)


@login_required
@permission_required('can_delete_customers')
def customer_delete(request, customer_id):
    """Delete customer"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        customer_name = customer.name
        customer.delete()
        messages.success(request, f'Customer "{customer_name}" deleted successfully!')
        return redirect('customers_list')
    
    return render(request, 'customer_delete.html', {'customer': customer})


# ============ ORDER VIEWS ============

from django.db.models import Prefetch
@login_required
@permission_required('can_view_orders')
def orders_list(request):
    """Display list of orders with filters and statistics"""
    from datetime import timedelta
    from django.utils import timezone
    from django.db.models import Q, Sum
    from decimal import Decimal, InvalidOperation
    
    # Get all orders initially
    orders = Order.objects.select_related('customer', 'created_by').filter(
        is_deleted=False
    ).order_by('-created_at')
    
    # GET FILTER PARAMETERS - DEFAULT TO 'today'
    date_filter = request.GET.get('date_range', 'today')
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    payment_filter = request.GET.get('payment', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    logistics_filter = request.GET.get('logistics_status', '')  # ✅ NEW: Logistics filter
    
    # Search filter
    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(customer_phone__icontains=search_query) |
            Q(customer_email__icontains=search_query)
        )
    
    # Status filter
    if status_filter:
        orders = orders.filter(order_status=status_filter)
    
    # Payment filter
    if payment_filter:
        orders = orders.filter(payment_status=payment_filter)
    
    # ✅ FIXED: LOGISTICS STATUS FILTER
    if logistics_filter == 'sent':
        # Show only orders successfully sent to NCM (must have an NCM ID)
        orders = orders.filter(ncm_order_id__isnull=False)
        
    elif logistics_filter == 'not_sent':
        # Show ALL orders that haven't been sent to NCM yet
        # REMOVED: Q(logistics='ncm') requirement
        # This now includes orders with logistics=NULL or no NCM ID
        orders = orders.filter(ncm_order_id__isnull=True)
    
    # DATE RANGE FILTER
    today = timezone.now().date()
    
    if date_filter == 'today':
        orders = orders.filter(created_at__date=today)
    elif date_filter == 'yesterday':
        yesterday = today - timedelta(days=1)
        orders = orders.filter(created_at__date=yesterday)
    elif date_filter == 'last_7_days':
        start = today - timedelta(days=7)
        orders = orders.filter(created_at__date__gte=start, created_at__date__lte=today)
    elif date_filter == 'last_30_days':
        start = today - timedelta(days=30)
        orders = orders.filter(created_at__date__gte=start, created_at__date__lte=today)
    elif date_filter == 'this_month':
        orders = orders.filter(created_at__year=today.year, created_at__month=today.month)
    elif date_filter == 'last_month':
        first_day_this_month = today.replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        first_day_last_month = last_day_last_month.replace(day=1)
        orders = orders.filter(created_at__date__gte=first_day_last_month, created_at__date__lte=last_day_last_month)
    elif date_filter == 'this_year':
        orders = orders.filter(created_at__year=today.year)
    elif date_filter == 'custom' and start_date and end_date:
        orders = orders.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
    elif date_filter == 'all':
        pass
    
    # FIX: Convert queryset to list and fix decimals
    orders_list = list(orders)
    for order in orders_list:
        try:
            order.discount_amount = order.discount_amount if order.discount_amount is not None else Decimal('0.00')
            order.shipping_charge = order.shipping_charge if order.shipping_charge is not None else Decimal('0.00')
            order.tax_percent = order.tax_percent if order.tax_percent is not None else Decimal('0.00')
            order.total_amount = order.total_amount if order.total_amount is not None else Decimal('0.00')
            order.partial_amount_paid = order.partial_amount_paid if order.partial_amount_paid is not None else Decimal('0.00')
            order.remaining_amount = order.remaining_amount if order.remaining_amount is not None else Decimal('0.00')
        except (InvalidOperation, ValueError, TypeError):
            order.discount_amount = Decimal('0.00')
            order.shipping_charge = Decimal('0.00')
            order.tax_percent = Decimal('0.00')
            order.total_amount = Decimal('0.00')
            order.partial_amount_paid = Decimal('0.00')
            order.remaining_amount = Decimal('0.00')
    
    # Statistics
    total_orders = len(orders_list)
    total_revenue = sum(o.total_amount for o in orders_list)
    pending_orders = len([o for o in orders_list if o.order_status == 'pending'])
    confirmed_orders = len([o for o in orders_list if o.order_status == 'confirmed'])
    dispatched_orders = len([o for o in orders_list if o.order_status == 'dispatched'])
    
    # Delivered today
    delivered_today = Order.objects.filter(
        order_status='delivered',
        delivered_at__date=today
    ).count()
    
    # Get first product name for each order
    order_products = {}
    for order in orders_list:
        try:
            first_item = order.items.first()
            if first_item:
                order_products[order.id] = first_item.product_name
            else:
                order_products[order.id] = "No products"
        except:
            order_products[order.id] = "No products"
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(orders_list, 25)
    page_number = request.GET.get('page')
    orders_page = paginator.get_page(page_number)
    
    context = {
        'orders': orders_page,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'pending_orders': pending_orders,
        'delivered_today': delivered_today,
        'confirmed_orders': confirmed_orders,
        'dispatched_orders': dispatched_orders,
        'search_query': search_query,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'date_filter': date_filter,
        'start_date': start_date,
        'end_date': end_date,
        'logistics_filter': logistics_filter,
        'order_products': order_products,
    }
    
    return render(request, 'orders_list.html', context)
@login_required
@permission_required('can_create_orders')
def order_create(request):
    """Create a new order with city management integration and custom product support"""
    if request.method == "POST":
        try:
            with transaction.atomic():
                customer_name = (request.POST.get("customer_name") or "").strip()
                customer_phone = (request.POST.get("customer_phone") or "").strip()
                customer_email = (request.POST.get("customer_email") or "").strip()
                # ✅ UPDATED: Get city from City model
                branch_city_name = (request.POST.get("branch_city") or "").strip()
                shipping_address = (request.POST.get("shipping_address") or "").strip()
                landmark = (request.POST.get("landmark") or "").strip()
                # ✅ Get in_out field from form (auto-detected)
                in_out = (request.POST.get("in_out") or "in").strip()

                created_by_id = request.POST.get("created_by")
                order_from = request.POST.get("order_from")
                order_status = request.POST.get("order_status") or "processing"
                payment_method = request.POST.get("payment_method") or "cod"
                payment_status = request.POST.get("payment_status") or "pending"

                discount_amount = Decimal(request.POST.get("discount") or "0")
                shipping_charge = Decimal(request.POST.get("shipping_charge") or "0")
                tax_percent = Decimal(request.POST.get("tax_percent") or "0")
                total_amount = Decimal(request.POST.get("total_amount") or "0")
                notes = request.POST.get("notes") or ""

                # ✅ GET PARTIAL PAYMENT DATA
                is_partial_payment = request.POST.get("is_partial_payment") == "true"
                partial_amount_paid = Decimal(request.POST.get("partial_amount_paid") or "0")
                remaining_amount = Decimal(request.POST.get("remaining_amount") or "0")

                # ✅ UPDATED: Added in_out to required fields check
                if not (customer_name and customer_phone and branch_city_name and shipping_address and created_by_id and in_out):
                    messages.error(request, "Please fill all required fields.")
                    return redirect("order_create")

                created_by = get_object_or_404(User, id=created_by_id)

                # ✅ Get or create city from City model
                city, city_created = City.objects.get_or_create(
                    name=branch_city_name,
                    defaults={
                        'valley_status': 'valley' if in_out.lower() == 'in' else 'out_valley',
                        'is_active': True
                    }
                )
                
                # Update city valley status if needed
                if not city_created and in_out.lower() == 'in' and city.valley_status != 'valley':
                    city.valley_status = 'valley'
                    city.save()
                elif not city_created and in_out.lower() == 'out' and city.valley_status != 'out_valley':
                    city.valley_status = 'out_valley'
                    city.save()

                customer, created = Customer.objects.get_or_create(
                    phone=customer_phone,
                    defaults={
                        "name": customer_name,
                        "email": customer_email or None,
                        "city": branch_city_name,
                        "address": shipping_address,
                        "landmark": landmark,
                    },
                )
                customer.name = customer_name
                customer.email = customer_email or None
                customer.city = branch_city_name
                customer.address = shipping_address
                customer.landmark = landmark
                customer.save()

                last_order = Order.objects.order_by("-id").first()
                if last_order and last_order.order_number.startswith("ORD"):
                    try:
                        n = int(last_order.order_number.replace("ORD", ""))
                    except ValueError:
                        n = last_order.id
                    order_number = f"ORD{n+1:06d}"
                else:
                    order_number = "ORD000001"

                order_items_json = request.POST.get("order_items") or "[]"
                cart = json.loads(order_items_json)

                if not cart:
                    messages.error(request, "No products in cart.")
                    return redirect("order_create")

                # ✅ SET PAYMENT STATUS BASED ON PARTIAL PAYMENT
                if is_partial_payment:
                    payment_status = "partial"

                # ✅ UPDATED: Use branch_city from City model, added in_out field
                order = Order.objects.create(
                    order_number=order_number,
                    created_by=created_by,
                    customer=customer,
                    customer_name=customer_name,
                    customer_phone=customer_phone,
                    customer_email=customer_email,
                    branch_city=branch_city_name,
                    in_out=in_out,
                    shipping_address=shipping_address,
                    landmark=landmark,
                    order_from=order_from,
                    order_status=order_status,
                    payment_method=payment_method,
                    payment_status=payment_status,
                    discount_amount=discount_amount,
                    shipping_charge=shipping_charge,
                    tax_percent=tax_percent,
                    total_amount=total_amount,
                    notes=notes,
                    # ✅ ADD PARTIAL PAYMENT FIELDS
                    is_partial_payment=is_partial_payment,
                    partial_amount_paid=partial_amount_paid if is_partial_payment else None,
                    remaining_amount=remaining_amount if is_partial_payment else None,
                )

                # CREATE ORDER ITEMS
                for item in cart:
                    product_id = int(item.get("id"))
                    var_id = item.get("varId")
                    qty = int(item.get("qty") or 1)
                    price = Decimal(str(item.get("price") or "0"))
                    sku = item.get("sku") or ""

                    product = get_object_or_404(Product, id=product_id)

                    variation = None
                    variation_name = None
                    if var_id:
                        variation = get_object_or_404(ProductVariation, id=int(var_id), product=product)
                        sku = variation.sku
                        variation_name = getattr(variation, 'variation_name', None) or variation.sku

                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        product_variation=variation,
                        product_name=product.name,
                        product_sku=sku,
                        variation_name=variation_name,
                        quantity=qty,
                        price=price,
                        total=price * qty,
                    )

                # ✅ CREATE ACTIVITY LOG WITH PARTIAL PAYMENT INFO
                description = f"Order #{order.order_number} was created with total amount रू {order.total_amount}"
                if is_partial_payment:
                    description += f" | Partial Payment: रू {partial_amount_paid} paid, रू {remaining_amount} remaining"
                
                OrderActivityLog.objects.create(
                    order=order,
                    action_type='created',
                    user=created_by,
                    description=description
                )

                # ✅ ADD CITY DETECTION LOG
                valley_status = "Valley" if in_out.lower() == 'in' else "Out Valley"
                OrderActivityLog.objects.create(
                    order=order,
                    action_type='city_detected',
                    user=created_by,
                    description=f'City "{branch_city_name}" detected as {valley_status}. IN/OUT set to {in_out.upper()}'
                )

                success_msg = f"Order {order.order_number} created successfully!"
                if is_partial_payment:
                    success_msg += f" | Partial payment: रू {partial_amount_paid} paid"
                
                messages.success(request, success_msg)
                return redirect("orders_list")

        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error creating order: {str(e)}")
            return redirect("order_create")

    # ✅ GET REQUEST - SHOW FORM
    users = User.objects.filter(is_active=True).order_by("username")
    recent_orders = Order.objects.filter(is_deleted=False).order_by("-created_at")[:6]
    
    # ✅ GET CITIES FROM DATABASE
    cities = City.objects.filter(is_active=True).order_by('name')
    
    # ✅ GET CATEGORIES FOR CUSTOM PRODUCT MODAL
    categories = Category.objects.all().order_by('name')
    
    return render(
        request,
        "order_create.html",
        {
            "users": users,
            "recent_orders": recent_orders,
            "cities": cities,
            "categories": categories,  # ✅ ADDED FOR CUSTOM PRODUCT FEATURE
        },
    )
@login_required
@permission_required('can_view_orders')
def order_detail(request, order_id):
    """View order details with partial payment info and status updates"""
    order = get_object_or_404(Order, id=order_id)
    order = fix_order_decimals(order)
    
    # ✅ ENSURE PARTIAL PAYMENT FIELDS ARE PROPERLY SET
    if order.payment_status == 'partial' and not order.is_partial_payment:
        order.is_partial_payment = True
        if order.partial_amount_paid is None:
            order.partial_amount_paid = Decimal('0.00')
        if order.remaining_amount is None:
            if order.total_amount and order.partial_amount_paid:
                order.remaining_amount = order.total_amount - order.partial_amount_paid
            else:
                order.remaining_amount = order.total_amount
        order.save()
    
    # Handle POST request for status updates
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            try:
                old_order_status = order.order_status
                old_payment_status = order.payment_status
                old_tracking = order.tracking_number or ''
                old_admin_notes = order.admin_notes or ''
                old_logistics = order.logistics or ''  # ✅ NEW: Track logistics changes
                
                # Get new values from form
                new_order_status = request.POST.get('order_status', order.order_status)
                new_payment_status = request.POST.get('payment_status', order.payment_status)
                new_tracking = request.POST.get('tracking_number', '').strip()
                new_admin_notes = request.POST.get('admin_notes', '').strip()
                new_logistics = request.POST.get('logistics', '').strip()  # ✅ NEW: Get logistics value
                
                # ✅ ENHANCED PARTIAL PAYMENT HANDLING
                # If changing TO partial payment status
                if old_payment_status != 'partial' and new_payment_status == 'partial':
                    order.is_partial_payment = True
                    # Initialize partial payment amounts if not set
                    if order.partial_amount_paid is None:
                        order.partial_amount_paid = Decimal('0.00')
                    if order.remaining_amount is None:
                        order.remaining_amount = order.total_amount
                    
                    # Create activity log for enabling partial payment
                    OrderActivityLog.objects.create(
                        order=order,
                        action_type='payment_changed',
                        user=request.user,
                        field_name='partial_payment_enabled',
                        old_value=old_payment_status,
                        new_value='partial',
                        description=f'Partial payment enabled. Payment status changed from "{old_payment_status}" to "partial"'
                    )
                
                # If changing FROM partial payment status
                elif old_payment_status == 'partial' and new_payment_status != 'partial':
                    order.is_partial_payment = False
                    
                    # Create activity log for clearing partial payment
                    OrderActivityLog.objects.create(
                        order=order,
                        action_type='payment_changed',
                        user=request.user,
                        field_name='partial_payment_cleared',
                        old_value='partial',
                        new_value=new_payment_status,
                        description=f'Partial payment cleared. Payment status changed from "partial" to "{new_payment_status}"'
                    )
                
                # Update order fields
                order.order_status = new_order_status
                order.payment_status = new_payment_status
                order.tracking_number = new_tracking
                order.admin_notes = new_admin_notes
                order.logistics = new_logistics  # ✅ NEW: Update logistics field
                
                # ✅ ADD IN/OUT FIELD UPDATE SUPPORT (if provided)
                new_in_out = request.POST.get('in_out')
                if new_in_out and new_in_out in ['in', 'out'] and order.in_out != new_in_out:
                    OrderActivityLog.objects.create(
                        order=order,
                        action_type='updated',
                        user=request.user,
                        field_name='in_out',
                        old_value=order.in_out,
                        new_value=new_in_out,
                        description=f'IN/OUT status changed from "{order.in_out}" to "{new_in_out}"'
                    )
                    order.in_out = new_in_out
                
                # Set delivered_at timestamp if status changed to delivered
                if order.order_status == 'delivered' and old_order_status != 'delivered':
                    order.delivered_at = timezone.now()
                
                order.save()
                
                # Create activity logs for changes
                changes_made = []
                
                # Order status changed
                if old_order_status != order.order_status:
                    OrderActivityLog.objects.create(
                        order=order,
                        action_type='status_changed',
                        user=request.user,
                        field_name='order_status',
                        old_value=old_order_status,
                        new_value=order.order_status,
                        description=f'Order status changed from "{old_order_status}" to "{order.order_status}"'
                    )
                    changes_made.append('Order Status')
                
                # Payment status changed
                if old_payment_status != order.payment_status:
                    OrderActivityLog.objects.create(
                        order=order,
                        action_type='payment_changed',
                        user=request.user,
                        field_name='payment_status',
                        old_value=old_payment_status,
                        new_value=order.payment_status,
                        description=f'Payment status changed from "{old_payment_status}" to "{order.payment_status}"'
                    )
                    changes_made.append('Payment Status')
                
                # ✅ NEW: Logistics changed
                if old_logistics != new_logistics:
                    logistics_display = {
                        'ncm': 'NCM',
                        'sundarijal': 'Sundarijal',
                        'express': 'Express',
                        'local': 'Local Delivery',
                        'other': 'Other',
                        '': 'None'
                    }
                    old_display = logistics_display.get(old_logistics, old_logistics or 'None')
                    new_display = logistics_display.get(new_logistics, new_logistics or 'None')
                    
                    OrderActivityLog.objects.create(
                        order=order,
                        action_type='updated',
                        user=request.user,
                        field_name='logistics',
                        old_value=old_logistics,
                        new_value=new_logistics,
                        description=f'Logistics provider changed from "{old_display}" to "{new_display}"'
                    )
                    changes_made.append('Logistics Provider')
                
                # Tracking number added or updated
                if old_tracking != new_tracking:
                    if old_tracking == '':
                        OrderActivityLog.objects.create(
                            order=order,
                            action_type='tracking_added',
                            user=request.user,
                            field_name='tracking_number',
                            new_value=new_tracking,
                            description=f'Tracking number added: {new_tracking}'
                        )
                        changes_made.append('Tracking Number Added')
                    else:
                        OrderActivityLog.objects.create(
                            order=order,
                            action_type='tracking_updated',
                            user=request.user,
                            field_name='tracking_number',
                            old_value=old_tracking,
                            new_value=new_tracking,
                            description=f'Tracking number updated from "{old_tracking}" to "{new_tracking}"'
                        )
                        changes_made.append('Tracking Number Updated')
                
                # Admin notes added or updated
                if old_admin_notes != new_admin_notes:
                    if old_admin_notes == '':
                        OrderActivityLog.objects.create(
                            order=order,
                            action_type='notes_added',
                            user=request.user,
                            field_name='admin_notes',
                            new_value=new_admin_notes[:100],
                            description='Admin notes added'
                        )
                        changes_made.append('Admin Notes Added')
                    else:
                        OrderActivityLog.objects.create(
                            order=order,
                            action_type='notes_updated',
                            user=request.user,
                            field_name='admin_notes',
                            old_value=old_admin_notes[:100],
                            new_value=new_admin_notes[:100],
                            description='Admin notes updated'
                        )
                        changes_made.append('Admin Notes Updated')
                
                if changes_made:
                    messages.success(request, f"✅ Order updated successfully! Changed: {', '.join(changes_made)}")
                else:
                    messages.info(request, "ℹ️ No changes were made to the order.")
                
                return redirect('order_detail', order_id=order.id)
                
            except Exception as e:
                messages.error(request, f"❌ Error updating order: {str(e)}")
                return redirect('order_detail', order_id=order.id)
    
    # GET request - display order details
    order_items = order.items.select_related('product', 'product_variation').all()
    activity_logs = order.activity_logs.select_related('user').order_by('-created_at')[:20]
    
    # Calculate subtotal
    subtotal = sum(item.total for item in order_items) or Decimal('0.00')
    
    # Calculate amounts
    after_discount = subtotal - (order.discount_amount or Decimal('0'))
    tax_amount = (after_discount * (order.tax_percent or Decimal('0'))) / 100
    
    # ✅ CALCULATE PARTIAL PAYMENT INFO
    is_partial_payment = order.is_partial_payment or order.payment_status == 'partial'
    partial_amount_paid = order.partial_amount_paid or Decimal('0.00')
    
    # Auto-calculate remaining amount if not set
    if is_partial_payment and order.remaining_amount is None:
        remaining_amount = (order.total_amount or Decimal('0.00')) - partial_amount_paid
        if remaining_amount < 0:
            remaining_amount = Decimal('0.00')
    else:
        remaining_amount = order.remaining_amount or Decimal('0.00')
    
    context = {
        'order': order,
        'order_items': order_items,
        'activity_logs': activity_logs,
        'subtotal': subtotal,
        'after_discount': after_discount,
        'tax_amount': tax_amount,
        
        # ✅ ENHANCED PARTIAL PAYMENT INFO
        'is_partial_payment': is_partial_payment,
        'partial_amount_paid': partial_amount_paid,
        'remaining_amount': remaining_amount,
        
        # ✅ CALCULATE PARTIAL PAYMENT PERCENTAGE FOR PROGRESS BAR
        'partial_payment_percentage': 0,
    }
    
    # Calculate percentage for progress bar
    if is_partial_payment and order.total_amount and order.total_amount > 0:
        try:
            percentage = (partial_amount_paid / order.total_amount) * 100
            context['partial_payment_percentage'] = min(100, max(0, float(percentage)))
        except:
            context['partial_payment_percentage'] = 0
    
    return render(request, 'order_detail.html', context)

@login_required
@permission_required('can_edit_orders')
def order_edit(request, order_id):
    """Edit an existing order with city management integration"""
    order = get_object_or_404(Order, id=order_id)
    order_items = order.items.all()

    if request.method == "POST":
        try:
            with transaction.atomic():
                # Store old values for activity log
                old_payment_method = order.payment_method
                old_is_partial = order.is_partial_payment
                old_partial_paid = order.partial_amount_paid or Decimal('0')
                old_remaining = order.remaining_amount or Decimal('0')
                old_city = order.branch_city
                old_in_out = order.in_out
                
                # Update basic fields
                order.customer_name = request.POST.get("customer_name", "").strip()
                order.customer_phone = request.POST.get("customer_phone", "").strip()
                order.customer_email = request.POST.get("customer_email", "").strip()
                # ✅ UPDATED: Get city from City model
                branch_city_name = request.POST.get("branch_city", "").strip()
                # ✅ Get in_out field from form (auto-detected)
                in_out = request.POST.get("in_out", "in").strip()
                order.shipping_address = request.POST.get("shipping_address", "").strip()
                order.landmark = request.POST.get("landmark", "").strip()
                
                created_by_id = request.POST.get("created_by")
                order.created_by = get_object_or_404(User, id=created_by_id)
                
                order.order_from = request.POST.get("order_from")
                order.order_status = request.POST.get("order_status")
                order.payment_method = request.POST.get("payment_method")
                
                order.discount_amount = Decimal(request.POST.get("discount") or "0")
                order.shipping_charge = Decimal(request.POST.get("shipping_charge") or "0")
                order.tax_percent = Decimal(request.POST.get("tax_percent") or "0")
                order.total_amount = Decimal(request.POST.get("total_amount") or "0")
                order.notes = request.POST.get("notes", "")

                # ✅ UPDATE PARTIAL PAYMENT DATA
                is_partial_payment = request.POST.get("is_partial_payment") == "true"
                partial_amount_paid = Decimal(request.POST.get("partial_amount_paid") or "0")
                remaining_amount = Decimal(request.POST.get("remaining_amount") or "0")
                
                order.is_partial_payment = is_partial_payment
                order.partial_amount_paid = partial_amount_paid if is_partial_payment else None
                order.remaining_amount = remaining_amount if is_partial_payment else None
                
                # ✅ UPDATE PAYMENT STATUS BASED ON PARTIAL PAYMENT
                if is_partial_payment:
                    if partial_amount_paid >= order.total_amount:
                        order.payment_status = "paid"
                    elif partial_amount_paid > 0:
                        order.payment_status = "partial"
                    else:
                        order.payment_status = "pending"

                # ✅ Get or create city from City model
                if branch_city_name:
                    city, city_created = City.objects.get_or_create(
                        name=branch_city_name,
                        defaults={
                            'valley_status': 'valley' if in_out.lower() == 'in' else 'out_valley',
                            'is_active': True
                        }
                    )
                    
                    # Update city valley status if needed
                    if not city_created and in_out.lower() == 'in' and city.valley_status != 'valley':
                        city.valley_status = 'valley'
                        city.save()
                    elif not city_created and in_out.lower() == 'out' and city.valley_status != 'out_valley':
                        city.valley_status = 'out_valley'
                        city.save()
                    
                    order.branch_city = branch_city_name
                
                order.in_out = in_out

                # Update customer
                if order.customer:
                    order.customer.name = order.customer_name
                    order.customer.email = order.customer_email or None
                    order.customer.phone = order.customer_phone
                    order.customer.city = order.branch_city
                    order.customer.address = order.shipping_address
                    order.customer.landmark = order.landmark
                    order.customer.save()

                # Update order items
                order.items.all().delete()
                
                order_items_json = request.POST.get("order_items") or "[]"
                cart = json.loads(order_items_json)

                if not cart:
                    messages.error(request, "No products in cart.")
                    return redirect("order_edit", order_id=order.id)

                for item in cart:
                    product_id = int(item.get("id"))
                    var_id = item.get("varId")
                    qty = int(item.get("qty") or 1)
                    price = Decimal(str(item.get("price") or "0"))
                    sku = item.get("sku") or ""

                    product = get_object_or_404(Product, id=product_id)
                    variation = None
                    variation_name = None
                    if var_id:
                        variation = get_object_or_404(ProductVariation, id=int(var_id))
                        sku = variation.sku
                        variation_name = getattr(variation, 'variation_name', None) or variation.sku

                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        product_variation=variation,
                        product_name=product.name,
                        product_sku=sku,
                        variation_name=variation_name,
                        quantity=qty,
                        price=price,
                        total=price * qty,
                    )

                order.save()

                # ✅ CREATE ACTIVITY LOG FOR PARTIAL PAYMENT CHANGES
                description = f"Order #{order.order_number} was updated"
                
                # Check if partial payment changed
                if is_partial_payment != old_is_partial:
                    if is_partial_payment:
                        description += f" | Changed to Partial Payment: रू {partial_amount_paid} paid, रू {remaining_amount} remaining"
                    else:
                        description += f" | Changed from Partial Payment to {order.payment_method.upper()}"
                elif is_partial_payment:
                    if partial_amount_paid != old_partial_paid:
                        description += f" | Partial Payment Updated: रू {partial_amount_paid} paid, रू {remaining_amount} remaining"
                
                # Check if city changed
                if old_city != branch_city_name:
                    description += f" | City changed from {old_city} to {branch_city_name}"
                
                # Check if IN/OUT changed
                if old_in_out != in_out:
                    description += f" | IN/OUT changed from {old_in_out.upper()} to {in_out.upper()}"
                
                OrderActivityLog.objects.create(
                    order=order,
                    action_type='updated',
                    user=request.user,
                    description=description
                )

                # Success message
                success_msg = f"✅ Order {order.order_number} updated successfully!"
                if is_partial_payment:
                    success_msg += f" | Partial payment: रू {partial_amount_paid} paid"
                
                messages.success(request, success_msg)
                return redirect("order_detail", order_id=order.id)

        except json.JSONDecodeError:
            messages.error(request, "Invalid cart data.")
            return redirect("order_edit", order_id=order.id)
        except Exception as e:
            messages.error(request, f"Error updating order: {str(e)}")
            print(f"Order update error: {e}")
            import traceback
            traceback.print_exc()
            return redirect("order_edit", order_id=order.id)

    # GET request - show form
    users = User.objects.filter(is_active=True).order_by("username")
    statuses = ["processing", "confirmed", "shipped", "delivered", "cancelled"]
    order_sources = ["website", "facebook", "instagram", "phone", "walk-in"]
    payment_methods = ["cod", "esewa", "khalti", "bank", "cash", "partial"]

    # Prepare JSON for initial client-side cart (safe and linter-friendly)
    initial_items_list = []
    for item in order_items:
        if item.product:
            initial_items_list.append({
                'id': item.product.id,
                'varId': item.product_variation.id if item.product_variation else None,
                'name': item.product_name,
                'sku': item.product_sku or '',
                'price': float(item.price),
                'qty': item.quantity,
            })
    initial_items_json = json.dumps(initial_items_list)

    # ✅ Get active cities for Branch/City select (from City management)
    cities = City.objects.filter(is_active=True).order_by('name')

    context = {
        "order": order,
        "order_items": order_items,
        "initial_items_json": initial_items_json,
        "users": users,
        "statuses": statuses,
        "order_sources": order_sources,
        "payment_methods": payment_methods,
        "recent_orders": Order.objects.all().order_by("-created_at")[:6],
        "cities": cities,
    }

    return render(request, "order_edit.html", context)

@login_required
@permission_required('can_delete_orders')
def order_delete(request, order_id):
    order = get_object_or_404(Order, id=order_id, created_by=request.user)

    if request.method == "POST":
        order_number = order.order_number

        # ❌ REMOVED STOCK RESTORATION - Stock was never reduced during order creation
        # Only orders with status="dispatched" have reduced stock
        # If you want to restore stock for dispatched orders, check status:
        
        if order.order_status == 'dispatched':
            # Restore stock for dispatched orders only
            for item in order.items.all():
                if item.product_variation:
                    item.product_variation.stock += item.quantity
                    if item.product_variation.stock > 0:
                        item.product_variation.status = 'active'
                    item.product_variation.save()
                elif item.product:
                    item.product.stock += item.quantity
                    if item.product.stock > 0:
                        item.product.stock_status = 'in_stock'
                    item.product.save()

        order.delete()
        messages.success(request, f"Order {order_number} deleted successfully!")
        return redirect("orders_list")

    return render(request, "order_delete.html", {"order": order})


@login_required
@permission_required('can_view_orders')
def orders_trash(request):
    """View trashed orders"""
    # ✅ Show all trashed orders (removed user filter)
    trashed_orders = Order.objects.filter(
        is_deleted=True
    ).order_by('-deleted_at')
    
    # Search functionality
    search_query = request.GET.get("search", "")
    if search_query:
        trashed_orders = trashed_orders.filter(
            Q(order_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(customer_phone__icontains=search_query)
        )
    
    # Status filter
    status_filter = request.GET.get("status", "")
    if status_filter:
        trashed_orders = trashed_orders.filter(order_status=status_filter)
    
    context = {
        "trashed_orders": trashed_orders,
        "search_query": search_query,
        "status_filter": status_filter,
    }
    
    return render(request, "orders_trash.html", context)


@login_required
@permission_required('can_delete_orders')
def order_move_to_trash(request, order_id):
    """Move order to trash (soft delete)"""
    # ✅ Removed user filter
    order = get_object_or_404(Order, id=order_id, is_deleted=False)
    
    if request.method == 'POST':
        order_number = order.order_number
        order.is_deleted = True
        order.deleted_at = timezone.now()
        order.save()
        
        # Log activity
        OrderActivityLog.objects.create(
            order=order,
            user=request.user,
            action_type='deleted',
            description=f'Order moved to trash by {request.user.username}'
        )
        
        messages.success(request, f'Order "{order_number}" moved to trash successfully!')
        return redirect('orders_list')
    
    return redirect('order_detail', order_id=order_id)


@login_required
@permission_required('can_delete_orders')
def order_restore(request, order_id):
    """Restore order from trash"""
    # ✅ Removed user filter
    order = get_object_or_404(Order, id=order_id, is_deleted=True)
    
    if request.method == 'POST':
        order_number = order.order_number
        order.is_deleted = False
        order.deleted_at = None
        order.save()
        
        # Log activity
        OrderActivityLog.objects.create(
            order=order,
            user=request.user,
            action_type='restored',
            description=f'Order restored from trash by {request.user.username}'
        )
        
        messages.success(request, f'Order "{order_number}" restored successfully!')
        return redirect('orders_trash')
    
    return redirect('orders_trash')


@login_required
@permission_required('can_delete_orders')
def order_permanent_delete(request, order_id):
    """Permanently delete order"""
    # ✅ Removed user filter
    order = get_object_or_404(Order, id=order_id, is_deleted=True)
    
    if request.method == 'POST':
        order_number = order.order_number
        order.delete()
        
        messages.success(request, f'Order "{order_number}" permanently deleted!')
        return redirect('orders_trash')
    
    return redirect('orders_trash')


@login_required
@permission_required('can_delete_orders')
def orders_trash_bulk_action(request):
    """Handle bulk actions on trashed orders"""
    if request.method == "POST":
        order_ids = request.POST.getlist("order_ids")
        action = request.POST.get("bulk_action")
        
        if not order_ids:
            messages.error(request, "No orders selected!")
            return redirect('orders_trash')
        
        try:
            # ✅ Removed user filter
            orders = Order.objects.filter(
                id__in=order_ids,
                is_deleted=True
            )
            count = orders.count()
            
            if count == 0:
                messages.error(request, "No valid orders found!")
                return redirect('orders_trash')
            
            if action == "restore":
                orders.update(is_deleted=False, deleted_at=None)
                
                # ✅ Log activity for each restored order
                for order in orders:
                    OrderActivityLog.objects.create(
                        order=order,
                        user=request.user,
                        action_type='restored',
                        description=f'Order restored from trash by {request.user.username}'
                    )
                
                messages.success(request, f"✅ {count} order(s) restored successfully!")
                
            elif action == "permanent_delete":
                orders.delete()
                messages.success(request, f"✅ {count} order(s) permanently deleted!")
                
            else:
                messages.error(request, "Invalid action selected!")
                
        except Exception as e:
            messages.error(request, f"Error performing bulk action: {str(e)}")
            
    return redirect('orders_trash')


@login_required
@permission_required('can_delete_orders')
def empty_orders_trash(request):
    """Empty all trashed orders"""
    if request.method == 'POST':
        # ✅ Removed user filter - empty ALL trashed orders
        trashed_orders = Order.objects.filter(is_deleted=True)
        count = trashed_orders.count()
        
        if count > 0:
            trashed_orders.delete()
            messages.success(request, f'✅ Trash emptied! {count} order(s) permanently deleted.')
        else:
            messages.info(request, 'Trash is already empty.')
        
        return redirect('orders_trash')
    
    return redirect('orders_trash')


@login_required
@permission_required('can_view_orders')
def order_invoice(request, order_id):
    order = get_object_or_404(Order, id=order_id, created_by=request.user)
    order_items = order.items.all()
    return render(request, "order_invoice.html", {
        "order": order,
        "order_items": order_items,
        "user": request.user,
    })


# API Endpoints for AJAX
@login_required
def api_get_customer(request, customer_id):
    """Get customer details via AJAX"""
    customer = get_object_or_404(Customer, id=customer_id)
    return JsonResponse({
        'name': customer.name,
        'email': customer.email,
        'phone': customer.phone,
        'address': customer.address,
        'city': customer.city,
        'state': customer.state,
        'postal_code': customer.postal_code or '',
    })


# ✅ SINGLE PRODUCT API (used by order edit/create modals)
@login_required
@require_http_methods(["GET"])
def api_get_product(request, product_id):
    """Return product details (non-variation) for POS modals"""
    try:
        # Role-aware access
        if request.user.is_superuser or getattr(request.user, 'role', None) in ['administrator', 'warehouse']:
            product = get_object_or_404(Product, id=product_id, is_deleted=False)
        else:
            product = get_object_or_404(Product, id=product_id, is_deleted=False, user=request.user)

        data = {
            'id': product.id,
            'name': product.name,
            'sku': product.sku if getattr(product, 'sku', None) else getattr(product, 'slug', ''),
            'price': str(product.price) if getattr(product, 'price', None) is not None else '0',
            'is_custom': bool(getattr(product, 'is_custom', False)),
            'category': {
                'id': product.category.id,
                'name': product.category.name
            } if getattr(product, 'category', None) else None,
            'description': product.description or '',
            'image': product.image.url if getattr(product, 'image', None) else None,
            'product_type': getattr(product, 'product_type', 'simple'),
            'stock': getattr(product, 'stock_quantity', getattr(product, 'stock', 0)),
        }

        return JsonResponse({'success': True, 'product': data})

    except Http404:
        return JsonResponse({'success': False, 'message': 'Product not found'}, status=404)
    except Exception as e:
        import traceback
        print(f"❌ Error in api_get_product: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def api_search_products(request):
    """
    Returns products for POS modal grid.
    Supports ?q= search. If q is empty, returns first 40 products.
    """
    try:
        q = (request.GET.get("q") or "").strip()

        # Role-aware visibility: administrators and warehouse users can see all products.
        # Other users (e.g., sales) will only see products they created.
        if request.user.is_superuser or getattr(request.user, 'role', None) in ['administrator', 'warehouse']:
            qs = Product.objects.filter(is_active=True, is_deleted=False).order_by("name")
        else:
            qs = Product.objects.filter(is_active=True, is_deleted=False, user=request.user).order_by("name")

        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(sku__icontains=q) | Q(slug__icontains=q))

        qs = qs[:40]

        data = []
        for p in qs:
            # ✅ Get stock - try different field names
            try:
                stock = p.stock_quantity if hasattr(p, 'stock_quantity') else (p.stock if hasattr(p, 'stock') else 0)
            except AttributeError:
                stock = 0
            
            # ✅ Get SKU - try sku field first, then slug as fallback
            try:
                sku = p.sku if hasattr(p, 'sku') and p.sku else p.slug
            except AttributeError:
                sku = p.slug

            data.append({
                "id": p.id,
                "name": p.name,
                "price": str(p.price) if p.price else "0",
                "stock": stock,
                "product_type": p.product_type,
                "image": p.image.url if p.image else None,
                "sku": sku,
            })

        return JsonResponse({
            "success": True,
            "products": data,
            "count": len(data)
        })
    
    except Exception as e:
        import traceback
        print(f"❌ Error in api_search_products: {str(e)}")
        traceback.print_exc()
        
        return JsonResponse({
            "success": False,
            "error": str(e),
            "products": []
        }, status=500)


# varialble product variations API
@login_required
@require_http_methods(["GET"])
def api_get_product_variations(request, product_id):
    """
    Returns variations for a variable product.
    JSON format matches your JS usage: data.variations[]
    """
    try:
        # Role-based product access
        if request.user.is_superuser or getattr(request.user, 'role', None) in ['administrator', 'warehouse']:
            product = get_object_or_404(Product, id=product_id, is_deleted=False)
        else:
            product = get_object_or_404(Product, id=product_id, is_deleted=False, user=request.user)

        # Check if variable product
        if product.product_type != "variable":
            return JsonResponse({
                "success": False,
                "message": "This product is not a variable product",
                "variations": []
            })

        # ✅ GET ALL ACTIVE VARIATIONS (ignore status field, only check is_active and stock)
        variations = (
            product.variations
            .filter(is_active=True)  # Only check is_active, ignore status
            .order_by("-stock", "sku")  # Show in-stock items first
        )

        if not variations.exists():
            return JsonResponse({
                "success": True,
                "variations": [],
                "total": 0,
                "message": "No active variations found for this product"
            })

        # Build response
        out = []
        for v in variations:
            # ✅ ONLY SHOW VARIATIONS WITH STOCK > 0
            if v.stock <= 0:
                continue
            
            # ✅ Get variation display name (prefer explicit variation_name, then name, then SKU)
            if hasattr(v, 'variation_name') and v.variation_name:
                variation_display = v.variation_name
            elif hasattr(v, 'name') and v.name:
                variation_display = v.name
            else:
                variation_display = v.sku
            
            out.append({
                "id": v.id,
                "sku": v.sku,
                "variation_name": variation_display,  # ✅ Added variation_name
                "price": str(v.price),
                "stock": v.stock,
                "is_active": v.is_active,
                "image": v.image.url if v.image else None,
                "category": {
                    "id": product.category.id,
                    "name": product.category.name
                } if getattr(product, 'category', None) else None,
            })

        return JsonResponse({
            "success": True,
            "variations": out,
            "total": len(out)
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error in api_get_product_variations: {str(e)}")
        traceback.print_exc()
        
        return JsonResponse({
            "success": False,
            "error": str(e),
            "variations": []
        }, status=500)

@login_required
@permission_required('can_export_data')
@require_http_methods(["GET"])
def export_orders_excel(request):
    """Export orders to Excel with ALL columns - like order details"""
    try:
        # ✅ Get all orders
        orders = Order.objects.filter(created_by=request.user).order_by('-created_at')
        
        # Apply filters from query parameters
        search_query = request.GET.get("search", "")
        status_filter = request.GET.get("status", "")
        payment_filter = request.GET.get("payment", "")
        
        if search_query:
            orders = orders.filter(
                Q(order_number__icontains=search_query) |
                Q(customer_name__icontains=search_query) |
                Q(customer_phone__icontains=search_query)
            )
        
        if status_filter:
            orders = orders.filter(order_status=status_filter)
        
        if payment_filter:
            orders = orders.filter(payment_status=payment_filter)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # ============= ROW 1: ALL HEADERS =============
        headers = [
            'Order ID', 'Order Number', 'Order Date', 'Order Status', 'Payment Status', 
            'Payment Method', 'Customer Name', 'Phone Number', 'Email Address', 
            'Shipping Address', 'Branch/City', 'Landmark', 'IN/OUT',  # ✅ UPDATED HEADERS
            'Product #', 'SKU', 'Product Name', 'Quantity', 'Unit Price', 'Total Price',
            'Grand Total'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # ✅ Get all items using RAW SQL to avoid decimal errors
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT order_id, product_sku, product_name, quantity, price
                FROM dashboard_orderitem
                WHERE order_id IN (
                    SELECT id FROM dashboard_order WHERE created_by_id = %s
                )
                ORDER BY order_id, id
            """, [request.user.id])
            
            all_items = cursor.fetchall()
        
        # Group items by order_id
        items_by_order = {}
        for item in all_items:
            order_id = item[0]
            if order_id not in items_by_order:
                items_by_order[order_id] = []
            items_by_order[order_id].append(item)
        
        # ============= WRITE ALL ROWS =============
        current_row = 2
        for order in orders:
            # Get items for this order
            order_items = items_by_order.get(order.id, [])
            
            # If no items, create one row with order info only
            if not order_items:
                row_data = [
                    order.id,
                    order.order_number,
                    order.created_at.strftime("%Y-%m-%d %H:%M"),
                    order.order_status.capitalize(),
                    order.payment_status.upper(),
                    order.payment_method.upper(),
                    order.customer_name or "N/A",
                    order.customer_phone or "N/A",
                    order.customer_email or "N/A",
                    order.shipping_address or "N/A",
                    order.branch_city or "N/A",  # ✅ UPDATED: branch_city instead of city
                    order.landmark or "N/A",
                    order.in_out.upper() if order.in_out else "IN",  # ✅ ADDED: in_out field
                    "",  # Product #
                    "",  # SKU
                    "",  # Product Name
                    "",  # Quantity
                    "",  # Unit Price
                    "",  # Total Price
                    float(order.total_amount)
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = center_alignment
                    if col_num in [20]:  # Adjusted for new column count
                        cell.number_format = '"रू "#,##0.00'
                
                current_row += 1
            else:
                # Write one row per item
                for idx, item in enumerate(order_items, 1):
                    product_sku = item[1] or "N/A"
                    product_name = item[2] or "N/A"
                    quantity = item[3] or 0
                    price = float(item[4]) if item[4] else 0.00
                    total_price = quantity * price
                    
                    row_data = [
                        order.id,
                        order.order_number,
                        order.created_at.strftime("%Y-%m-%d %H:%M"),
                        order.order_status.capitalize(),
                        order.payment_status.upper(),
                        order.payment_method.upper(),
                        order.customer_name or "N/A",
                        order.customer_phone or "N/A",
                        order.customer_email or "N/A",
                        order.shipping_address or "N/A",
                        order.branch_city or "N/A",  # ✅ UPDATED: branch_city instead of city
                        order.landmark or "N/A",
                        order.in_out.upper() if order.in_out else "IN",  # ✅ ADDED: in_out field
                        idx,  # Product #
                        product_sku,  # SKU
                        product_name,  # Product Name
                        quantity,  # Quantity
                        price,  # Unit Price
                        total_price,  # Total Price
                        float(order.total_amount)  # Grand Total
                    ]
                    
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.value = value
                        cell.border = border
                        cell.alignment = center_alignment
                        if col_num in [18, 19, 20]:  # Adjusted for new column count
                            cell.number_format = '"रू "#,##0.00'
                    
                    current_row += 1
        
        # Adjust column widths (added one more for IN/OUT)
        column_widths = [10, 15, 18, 12, 12, 12, 18, 15, 15, 20, 12, 15, 8, 8, 10, 20, 10, 12, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Create response
        filename = f"Orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error exporting orders: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error exporting orders: {str(e)}", status=500)
    
@login_required
@permission_required('can_export_data')
@require_http_methods(["GET"])
def export_order_details(request, order_id):
    """Export specific order with item details to Excel - ALL IN ONE ROW"""
    try:
        order = get_object_or_404(Order, id=order_id, created_by=request.user)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Order {order.order_number}"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # ============= ROW 1: ALL HEADERS IN ONE ROW =============
        headers = [
            'Order ID', 'Order Number', 'Order Date', 'Order Status', 'Payment Status', 
            'Payment Method', 'Customer Name', 'Phone Number', 'Email Address', 
            'Shipping Address', 'Branch/City', 'Landmark', 'IN/OUT',  # ✅ UPDATED HEADERS
            'Product #', 'SKU', 'Product Name', 'Quantity', 'Unit Price', 'Total Price',
            'Grand Total'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # ✅ Get items using RAW SQL
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT product_sku, product_name, quantity, price
                FROM dashboard_orderitem
                WHERE order_id = %s
            """, [order.id])
            
            items = cursor.fetchall()
        
        # ============= ROWS 2+: ONE ROW PER ITEM (WITH ALL DATA) =============
        current_row = 2
        for idx, item in enumerate(items, 1):
            product_sku = item[0] or "N/A"
            product_name = item[1] or "N/A"
            quantity = item[2] or 0
            price = float(item[3]) if item[3] else 0.00
            total_price = quantity * price
            
            # All data in one row - ✅ UPDATED FIELDS
            row_data = [
                order.id,                                          # Order ID
                order.order_number,                               # Order Number
                order.created_at.strftime("%Y-%m-%d %H:%M"),     # Order Date
                order.order_status.capitalize(),                  # Order Status
                order.payment_status.upper(),                     # Payment Status
                order.payment_method.upper(),                     # Payment Method
                order.customer_name or "N/A",                     # Customer Name
                order.customer_phone or "N/A",                    # Phone Number
                order.customer_email or "N/A",                    # Email Address
                order.shipping_address or "N/A",                  # Shipping Address
                order.branch_city or "N/A",                       # ✅ UPDATED: Branch/City
                order.landmark or "N/A",                          # Landmark
                order.in_out.upper() if order.in_out else "IN",  # ✅ ADDED: IN/OUT
                idx,                                              # Product #
                product_sku,                                      # SKU
                product_name,                                     # Product Name
                quantity,                                         # Quantity
                price,                                            # Unit Price
                total_price,                                      # Total Price
                float(order.total_amount)                         # Grand Total
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = center_alignment
                
                # Format price columns - ✅ ADJUSTED COLUMN NUMBERS
                if col_num in [18, 19, 20]:  # Unit Price, Total Price, Grand Total (adjusted for new columns)
                    cell.number_format = '"रू "#,##0.00'
            
            current_row += 1
        
        # Adjust column widths - ✅ ADDED ONE MORE COLUMN FOR IN/OUT
        column_widths = [10, 15, 18, 12, 12, 12, 18, 15, 15, 20, 12, 15, 8, 8, 10, 20, 10, 12, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Create response
        filename = f"Order_{order.order_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Error exporting order: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error exporting order: {str(e)}", status=500)

# new

@login_required
@require_POST
def product_gallery_upload(request, product_id):
    product = get_object_or_404(Product, pk=product_id, user=request.user)

    files = request.FILES.getlist("images")
    if not files:
        messages.error(request, "No files selected.")
        return redirect("product_detail", product_id=product_id)

    for f in files:
        ProductImage.objects.create(product=product, image=f)

    messages.success(request, f"{len(files)} image(s) uploaded.")
    return redirect("product_detail", product_id=product_id)


@login_required
@require_POST
def delete_product_image(request, image_id):
    img = get_object_or_404(ProductImage, pk=image_id)
    product_id = img.product_id

    if img.product.user != request.user:
        messages.error(request, "Permission denied.")
        return redirect("product_detail", product_id=product_id)

    img.delete()
    messages.success(request, "Image deleted.")
    return redirect("product_detail", product_id=product_id)


@login_required
@require_POST
@permission_required('can_create_products')
def variation_create(request, product_id):
    product = get_object_or_404(Product, pk=product_id, user=request.user)

    sku = request.POST.get("sku", "").strip()
    price = request.POST.get("price", "").strip()
    stock = request.POST.get("stock", "").strip()
    status = request.POST.get("status", "active").strip()

    if not sku or price == "" or stock == "":
        messages.error(request, "SKU, price and stock are required.")
        return redirect("product_detail", product_id=product_id)

    try:
        price_val = float(price)
        stock_val = int(stock)
    except ValueError:
        messages.error(request, "Invalid price/stock.")
        return redirect("product_detail", product_id=product_id)

    variation = ProductVariation.objects.create(
        product=product,
        sku=sku,
        price=price_val,
        stock=stock_val,
        status=status,
        image=request.FILES.get("image") if "image" in request.FILES else None,
    )

    messages.success(request, f'Variation "{variation.sku}" created.')
    return redirect("product_detail", product_id=product_id)


@login_required
@require_POST
@permission_required('can_edit_products')
def variation_update(request, variation_id):
    variation = get_object_or_404(ProductVariation, pk=variation_id)

    if variation.product.user != request.user:
        messages.error(request, "Permission denied.")
        return redirect("product_detail", product_id=variation.product_id)

    variation_name = request.POST.get("variation_name", "").strip()
    sku = request.POST.get("sku", "").strip()
    price = request.POST.get("price", "").strip()
    stock = request.POST.get("stock", "").strip()
    status = request.POST.get("status", "active").strip()

    if not variation_name or not sku or price == "" or stock == "":
        messages.error(request, "Variation name, SKU, price and stock are required.")
        return redirect("product_detail", product_id=variation.product_id)

    try:
        variation.price = float(price)
        variation.stock = int(stock)
    except ValueError:
        messages.error(request, "Invalid price/stock.")
        return redirect("product_detail", product_id=variation.product_id)

    variation.variation_name = variation_name
    variation.sku = sku
    variation.status = status

    if "image" in request.FILES:
        variation.image = request.FILES["image"]

    variation.save()
    messages.success(request, f'Variation "{variation.sku}" updated.')
    return redirect("product_detail", product_id=variation.product_id)


@login_required
@require_POST
def variation_delete(request, variation_id):
    variation = get_object_or_404(ProductVariation, pk=variation_id)
    product_id = variation.product_id

    if variation.product.user != request.user:
        messages.error(request, "Permission denied.")
        return redirect("product_detail", product_id=product_id)

    sku = variation.sku
    variation.delete()
    messages.success(request, f'Variation "{sku}" deleted.')
    return redirect("product_detail", product_id=product_id)

@login_required
@permission_required('can_delete_orders')
def orders_bulk_action(request):
    """Handle bulk actions on orders including NCM bulk sending"""
    if request.method == 'POST':
        order_ids = request.POST.getlist('order_ids')
        action = request.POST.get('bulk_action')
        
        if not order_ids:
            messages.error(request, 'No orders selected!')
            return redirect('orders_list')
        
        try:
            # ✅ Removed user filter - show all orders
            orders = Order.objects.filter(
                id__in=order_ids, 
                is_deleted=False
            )
            count = orders.count()
            
            if count == 0:
                messages.error(request, "No valid orders found!")
                return redirect('orders_list')
            
            # ✅ NEW: HANDLE SEND TO NCM ACTION
            if action == 'send_to_ncm':
                return bulk_send_to_ncm(request, orders)
            
            elif action == 'delete':
                # ✅ SOFT DELETE - Move to trash instead of permanent delete
                # Only restore stock for dispatched orders
                for order in orders:
                    if order.order_status == 'dispatched':
                        for item in order.items.all():
                            if item.product_variation:
                                item.product_variation.stock += item.quantity
                                if item.product_variation.stock > 0:
                                    item.product_variation.status = 'active'
                                item.product_variation.save()
                            elif item.product:
                                item.product.stock += item.quantity
                                if item.product.stock > 0:
                                    item.product.stock_status = 'in_stock'
                                item.product.save()
                    
                    # Log activity
                    OrderActivityLog.objects.create(
                        order=order,
                        user=request.user,
                        action_type='deleted',
                        description=f'Order moved to trash by {request.user.username}'
                    )
                
                # ✅ Soft delete instead of permanent delete
                orders.update(is_deleted=True, deleted_at=timezone.now())
                messages.success(request, f'✅ {count} order(s) moved to trash successfully!')
                
            elif action == 'mark_delivered':
                orders.update(order_status='delivered')
                
                # Log activity for each order
                for order in orders:
                    OrderActivityLog.objects.create(
                        order=order,
                        user=request.user,
                        action_type='status_changed',
                        description=f'Order status changed to Delivered by {request.user.username}'
                    )
                
                messages.success(request, f'✅ {count} order(s) marked as delivered!')
                
            elif action == 'mark_cancelled':
                orders.update(order_status='cancelled')
                
                # Log activity for each order
                for order in orders:
                    OrderActivityLog.objects.create(
                        order=order,
                        user=request.user,
                        action_type='status_changed',
                        description=f'Order status changed to Cancelled by {request.user.username}'
                    )
                
                messages.success(request, f'✅ {count} order(s) marked as cancelled!')
                
            elif action == 'mark_processing':
                orders.update(order_status='processing')
                
                # Log activity for each order
                for order in orders:
                    OrderActivityLog.objects.create(
                        order=order,
                        user=request.user,
                        action_type='status_changed',
                        description=f'Order status changed to Processing by {request.user.username}'
                    )
                
                messages.success(request, f'✅ {count} order(s) marked as processing!')
            
            elif action == 'mark_shipped':
                orders.update(order_status='shipped')
                
                # Log activity for each order
                for order in orders:
                    OrderActivityLog.objects.create(
                        order=order,
                        user=request.user,
                        action_type='status_changed',
                        description=f'Order status changed to Shipped by {request.user.username}'
                    )
                
                messages.success(request, f'✅ {count} order(s) marked as shipped!')
            
            elif action == 'mark_paid':
                orders.update(payment_status='paid')
                
                # Log activity for each order
                for order in orders:
                    OrderActivityLog.objects.create(
                        order=order,
                        user=request.user,
                        action_type='payment_changed',
                        description=f'Payment status changed to Paid by {request.user.username}'
                    )
                
                messages.success(request, f'✅ {count} order(s) marked as paid!')
            
            elif action == 'mark_pending':
                orders.update(payment_status='pending')
                
                # Log activity for each order
                for order in orders:
                    OrderActivityLog.objects.create(
                        order=order,
                        user=request.user,
                        action_type='payment_changed',
                        description=f'Payment status changed to Pending by {request.user.username}'
                    )
                
                messages.success(request, f'✅ {count} order(s) marked as pending payment!')
                
            else:
                messages.error(request, 'Invalid action selected!')
                
        except Exception as e:
            messages.error(request, f'Error performing bulk action: {str(e)}')
    
    return redirect('orders_list')


@login_required
@permission_required('can_delete_orders')
def orders_bulk_ncm_send(request):
    """
    Handle Bulk Sending to NCM Logistics
    """
    if request.method != 'POST':
        messages.error(request, '❌ Invalid request method')
        return redirect('orders_list')
    
    # 1. Get Form Data
    order_ids = request.POST.getlist('order_ids')
    from_branch = request.POST.get('from_branch', 'TINKUNE')
    delivery_type = request.POST.get('delivery_type', 'Door2Door')
    
    # Handle auto_set_logistics checkbox
    auto_set_logistics = request.POST.get('auto_set_logistics') == 'on'
    
    # Handle weight (default to 1.0 if invalid)
    try:
        default_weight = float(request.POST.get('default_weight', '1.0'))
    except (ValueError, TypeError):
        default_weight = 1.0

    if not order_ids:
        messages.error(request, '❌ No orders selected for NCM dispatch!')
        return redirect('orders_list')

    # 2. Get Orders
    orders = Order.objects.filter(id__in=order_ids, is_deleted=False)
    count = orders.count()

    if count == 0:
        messages.error(request, '❌ No valid active orders found matching selection.')
        return redirect('orders_list')

    # 3. Auto-set Logistics Field (if checked)
    if auto_set_logistics:
        orders.update(logistics='ncm')

    # 4. Processing Variables
    success_count = 0
    skip_count = 0
    error_count = 0
    error_details = []

    # 5. Iterate and Send
    for order in orders:
        result = send_single_order_to_ncm(
            request=request, 
            order=order, 
            from_branch=from_branch, 
            delivery_type=delivery_type, 
            default_weight=default_weight
        )

        if result['status'] == 'success':
            success_count += 1
        elif result['status'] == 'skipped':
            skip_count += 1
        else:
            error_count += 1
            # Capture the specific error message for the first few failures
            if len(error_details) < 3: 
                error_details.append(f"{order.order_number}: {result['message']}")

    # 6. Final Feedback
    if success_count > 0:
        messages.success(request, f'✅ Successfully sent {success_count} order(s) to NCM.')
    
    if skip_count > 0:
        messages.warning(request, f'⚠️ Skipped {skip_count} order(s) (Already sent or missing info).')
        
    if error_count > 0:
        messages.error(request, f'❌ Failed to send {error_count} order(s).')
        # Show specific API errors
        for err in error_details:
            messages.error(request, f"Error: {err}")

    return redirect('orders_list')

def send_single_order_to_ncm(request, order, from_branch='TINKUNE', delivery_type='Door2Door', default_weight=1.0):
    """
    Helper function to send a single order to NCM API.
    """
    try:
        # Check if already has NCM ID
        if order.ncm_order_id:
            return {'status': 'skipped', 'message': 'Already has NCM ID'}

        # 1. Get API Credentials from Settings
        # Ensure NCM_API_BASE_URL does NOT have a trailing slash
        base_url = getattr(settings, 'NCM_API_BASE_URL', '').rstrip('/')
        api_key = getattr(settings, 'NCM_API_KEY', '')

        if not base_url or not api_key:
            return {'status': 'error', 'message': 'NCM configuration missing in settings.py'}

        # Construct Endpoint (use /order/create not /ordercreate)
        api_url = f"{base_url}/order/create"

        # 2. Prepare Data
        # Ensure product name isn't None
        product_name = "General Item"
        first_item = order.items.first()
        if first_item:
            product_name = first_item.product_name or "General Item"

        # Calculate Weight
        weight = default_weight
        # If your order model has a weight field, use it
        if hasattr(order, 'weight') and order.weight:
             weight = float(order.weight)

        # Clean phone number (remove non-digits)
        phone = ''.join(filter(str.isdigit, str(order.customer_phone or "")))
        
        payload = {
            "name": str(order.customer_name or "").strip(),
            "phone": phone,
            "phone2": "", # Optional
            "cod_charge": float(order.total_amount or 0),
            "address": str(order.shipping_address or "").strip(),
            "fbranch": from_branch,
            "branch": str(order.branch_city or "KATHMANDU").upper(),
            "package": str(product_name)[:100], # Limit length
            "vrefid": str(order.order_number),
            "instruction": str(order.notes or "")[:100],
            "deliverytype": delivery_type,
            "weight": weight
        }

        # 3. Send Request
        headers = {
            'Authorization': f'Token {api_key}',
            'Content-Type': 'application/json'
        }

        response = requests.post(api_url, json=payload, headers=headers, timeout=15)

        # 4. Handle Response
        if response.status_code == 200:
            resp_data = response.json()
            
            # Check NCM specific success message
            if resp_data.get('Message') == 'Order Successfully Created':
                # Save NCM ID to Order
                order.ncm_order_id = str(resp_data.get('orderid'))
                order.logistics = 'ncm' # Ensure logistics is set
                order.save()

                # Log Activity
                OrderActivityLog.objects.create(
                    order=order,
                    user=request.user,
                    action_type='updated',
                    description=f"Sent to NCM. NCM ID: {order.ncm_order_id}"
                )
                return {'status': 'success', 'message': 'Sent successfully'}
            else:
                # API returned 200 but with an internal error message
                return {'status': 'error', 'message': str(resp_data)}
        
        elif response.status_code == 404:
            # 404 means the URL is wrong OR the Resource ID is wrong. 
            # Since we are creating, it's likely the URL.
            print(f"DEBUG: NCM 404 Error. Attempted URL: {api_url}")
            return {'status': 'error', 'message': f'API Endpoint 404. Checked URL: {api_url}'}
            
        else:
            return {'status': 'error', 'message': f'HTTP Error {response.status_code}: {response.text}'}

    except Exception as e:
        return {'status': 'error', 'message': str(e)}
# ==================== DISPATCH MANAGEMENT VIEWS ====================

@login_required
@permission_required('can_view_dispatch')
def dispatch_management(request):
    """Main dispatch scanning page"""
    if request.method == 'POST':
        order_ids_str = request.POST.get('order_ids', '').strip()
        set_status = request.POST.get('set_status')
        logistics = request.POST.get('logistics')
        
        if not order_ids_str:
            messages.error(request, 'Please scan at least one order ID.')
            return redirect('dispatch_management')
        
        if not set_status or not logistics:
            messages.error(request, 'Please select both status and logistics.')
            return redirect('dispatch_management')
        
        # Parse order IDs (comma-separated or newline-separated)
        order_ids_raw = order_ids_str.replace('\n', ',').replace('\r', '').split(',')
        order_ids = [oid.strip() for oid in order_ids_raw if oid.strip()]
        
        if not order_ids:
            messages.error(request, 'No valid order IDs found.')
            return redirect('dispatch_management')
        
        # Check for duplicates
        if len(order_ids) != len(set(order_ids)):
            messages.error(request, 'Duplicate order IDs detected. Please remove duplicates.')
            return redirect('dispatch_management')
        
        try:
            with transaction.atomic():
                # Generate batch number
                batch_number = f"DISPATCH-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
                
                # Create dispatch
                dispatch = Dispatch.objects.create(
                    batch_number=batch_number,
                    logistics=logistics,
                    status=set_status,
                    total_orders=len(order_ids),
                    created_by=request.user
                )
                
                # Create dispatch items and update orders
                updated_count = 0
                not_found = []
                stock_warnings = []
                
                for order_id in order_ids:
                    # Create dispatch item
                    DispatchItem.objects.create(
                        dispatch=dispatch,
                        scanned_order_id=order_id
                    )
                    
                    # Try to find and update order by order_number or barcode
                    order = Order.objects.filter(
                        Q(order_number=order_id) | Q(barcode=order_id)
                    ).first()
                    
                    if order:
                        # Check if already dispatched
                        if order.order_status == 'dispatched':
                            messages.warning(request, f'⚠️ Order {order_id} already dispatched')
                            continue
                        
                        # Reduce stock if status is "dispatched"
                        if set_status == 'dispatched':
                            for item in order.items.select_related('product', 'product_variation'):
                                product = item.product
                                variation = item.product_variation
                                quantity = item.quantity
                                
                                if variation:
                                    # Reduce variation stock
                                    if variation.stock >= quantity:
                                        variation.stock -= quantity
                                        if variation.stock == 0:
                                            variation.status = 'out_of_stock'
                                        variation.save()
                                    else:
                                        stock_warnings.append(
                                            f"⚠️ {variation.sku}: Need {quantity}, Available {variation.stock}"
                                        )
                                        variation.stock = max(0, variation.stock - quantity)
                                        variation.status = 'out_of_stock'
                                        variation.save()
                                else:
                                    # Reduce main product stock
                                    if product and product.stock >= quantity:
                                        product.stock -= quantity
                                        if product.stock == 0:
                                            product.stock_status = 'out_of_stock'
                                        elif product.stock <= 10:
                                            product.stock_status = 'low_stock'
                                        product.save()
                                    elif product:
                                        stock_warnings.append(
                                            f"⚠️ {product.name}: Need {quantity}, Available {product.stock}"
                                        )
                                        product.stock = max(0, product.stock - quantity)
                                        product.stock_status = 'out_of_stock'
                                        product.save()
                        
                        # Update order fields
                        order.order_status = set_status
                        order.logistics_provider = logistics
                        order.dispatch_date = timezone.now()
                        order.save()
                        
                        # Link dispatch item to order
                        DispatchItem.objects.filter(
                            dispatch=dispatch,
                            scanned_order_id=order_id
                        ).update(order=order)
                        
                        # Create activity log
                        OrderActivityLog.objects.create(
                            order=order,
                            action_type='status_changed',
                            user=request.user,
                            field_name='order_status',
                            old_value=order.order_status,
                            new_value=set_status,
                            description=f'Order dispatched via batch {batch_number} with {logistics}'
                        )
                        
                        updated_count += 1
                    else:
                        not_found.append(order_id)
                
                # Show stock warnings
                for warning in stock_warnings[:3]:
                    messages.warning(request, warning)
                if len(stock_warnings) > 3:
                    messages.warning(request, f'...and {len(stock_warnings) - 3} more stock warnings')
                
                # Success message
                if updated_count == len(order_ids):
                    messages.success(
                        request,
                        f'✅ Successfully dispatched {updated_count} orders! Batch: {batch_number}'
                    )
                else:
                    messages.warning(
                        request,
                        f'⚠️ Dispatched {updated_count}/{len(order_ids)} orders. '
                        f'{len(not_found)} order(s) not found: {", ".join(not_found)}'
                    )
                
                return redirect('dispatch_detail', pk=dispatch.pk)
                
        except Exception as e:
            messages.error(request, f'Error creating dispatch: {str(e)}')
            import traceback
            traceback.print_exc()
            return redirect('dispatch_management')
    
    return render(request, 'dispatch_management.html')


@login_required
@permission_required('can_view_dispatch')
def dispatch_list(request):
    """List all dispatches (not trashed)"""
    dispatches = Dispatch.objects.filter(is_deleted=False).prefetch_related('items').order_by('-created_at')
    
    # Filters
    logistics_filter = request.GET.get('logistics')
    status_filter = request.GET.get('status')
    search = request.GET.get('search')
    
    if logistics_filter:
        dispatches = dispatches.filter(logistics=logistics_filter)
    
    if status_filter:
        dispatches = dispatches.filter(status=status_filter)
    
    if search:
        dispatches = dispatches.filter(
            Q(batch_number__icontains=search) |
            Q(items__scanned_order_id__icontains=search)
        ).distinct()
    
    context = {
        'dispatches': dispatches,
        'logistics_choices': Dispatch.LOGISTICS_CHOICES,
        'status_choices': Dispatch.STATUS_CHOICES,
        'search': search,
        'logistics_filter': logistics_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'dispatch_list.html', context)


@login_required
@permission_required('can_view_dispatch')
def dispatch_detail(request, pk):
    """View single dispatch details"""
    dispatch = get_object_or_404(
        Dispatch.objects.prefetch_related('items__order'),
        pk=pk,
        is_deleted=False
    )
    
    context = {
        'dispatch': dispatch,
    }
    
    return render(request, 'dispatch_detail.html', context)


# ==================== DISPATCH TRASH MANAGEMENT ====================

@login_required
@permission_required('can_delete_dispatch')
def dispatch_move_to_trash(request, pk):
    """Move dispatch to trash (soft delete)"""
    dispatch = get_object_or_404(Dispatch, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        batch_number = dispatch.batch_number
        dispatch.is_deleted = True
        dispatch.deleted_by = request.user
        dispatch.deleted_at = timezone.now()
        dispatch.save()
        
        messages.success(request, f'Dispatch "{batch_number}" moved to trash successfully!')
        return redirect('dispatch_list')
    
    return redirect('dispatch_detail', pk=pk)


@login_required
@permission_required('can_view_dispatch')
def dispatch_trash(request):
    """View trashed dispatches"""
    trashed_dispatches = Dispatch.objects.filter(
        is_deleted=True
    ).select_related('created_by', 'deleted_by').prefetch_related('items').order_by('-deleted_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        trashed_dispatches = trashed_dispatches.filter(
            Q(batch_number__icontains=search_query) |
            Q(items__scanned_order_id__icontains=search_query)
        ).distinct()
    
    # Logistics filter
    logistics_filter = request.GET.get('logistics', '')
    if logistics_filter:
        trashed_dispatches = trashed_dispatches.filter(logistics=logistics_filter)
    
    context = {
        'trashed_dispatches': trashed_dispatches,
        'search_query': search_query,
        'logistics_filter': logistics_filter,
        'logistics_choices': Dispatch.LOGISTICS_CHOICES,
    }
    
    return render(request, 'dispatch_trash.html', context)


@login_required
@permission_required('can_delete_dispatch')
def dispatch_restore(request, pk):
    """Restore dispatch from trash"""
    dispatch = get_object_or_404(Dispatch, pk=pk, is_deleted=True)
    
    if request.method == 'POST':
        batch_number = dispatch.batch_number
        dispatch.is_deleted = False
        dispatch.deleted_by = None
        dispatch.deleted_at = None
        dispatch.save()
        
        messages.success(request, f'Dispatch "{batch_number}" restored successfully!')
        return redirect('dispatch_trash')
    
    return redirect('dispatch_trash')


@login_required
@permission_required('can_delete_dispatch')
def dispatch_permanent_delete(request, pk):
    """Permanently delete dispatch"""
    dispatch = get_object_or_404(Dispatch, pk=pk, is_deleted=True)
    
    if request.method == 'POST':
        batch_number = dispatch.batch_number
        dispatch.delete()
        
        messages.success(request, f'Dispatch "{batch_number}" permanently deleted!')
        return redirect('dispatch_trash')
    
    return redirect('dispatch_trash')


@login_required
@permission_required('can_delete_dispatch')
def dispatch_trash_bulk_action(request):
    """Handle bulk actions on trashed dispatches"""
    if request.method == 'POST':
        dispatch_ids = request.POST.getlist('dispatch_ids')
        action = request.POST.get('bulk_action')
        
        if not dispatch_ids:
            messages.error(request, 'No dispatches selected!')
            return redirect('dispatch_trash')
        
        try:
            dispatches = Dispatch.objects.filter(
                id__in=dispatch_ids,
                is_deleted=True
            )
            count = dispatches.count()
            
            if count == 0:
                messages.error(request, 'No valid dispatches found!')
                return redirect('dispatch_trash')
            
            if action == 'restore':
                dispatches.update(is_deleted=False, deleted_by=None, deleted_at=None)
                messages.success(request, f'✅ {count} dispatch(es) restored successfully!')
                
            elif action == 'permanent_delete':
                dispatches.delete()
                messages.success(request, f'✅ {count} dispatch(es) permanently deleted!')
                
            else:
                messages.error(request, 'Invalid action selected!')
                
        except Exception as e:
            messages.error(request, f'Error performing bulk action: {str(e)}')
    
    return redirect('dispatch_trash')


@login_required
@permission_required('can_delete_dispatch')
def empty_dispatch_trash(request):
    """Empty all trashed dispatches"""
    if request.method == 'POST':
        trashed_dispatches = Dispatch.objects.filter(is_deleted=True)
        count = trashed_dispatches.count()
        
        if count > 0:
            trashed_dispatches.delete()
            messages.success(request, f'✅ Trash emptied! {count} dispatch(es) permanently deleted.')
        else:
            messages.info(request, 'Trash is already empty.')
        
        return redirect('dispatch_trash')
    
    return redirect('dispatch_trash')

@login_required
@permission_required('can_view_dispatch')
def dispatch_list(request):
    """List all dispatches (not trashed)"""
    dispatches = Dispatch.objects.filter(is_deleted=False).prefetch_related('items').order_by('-created_at')
    
    # Filters
    logistics_filter = request.GET.get('logistics')
    status_filter = request.GET.get('status')
    search = request.GET.get('search')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if logistics_filter:
        dispatches = dispatches.filter(logistics=logistics_filter)
    
    if status_filter:
        dispatches = dispatches.filter(status=status_filter)
    
    if search:
        dispatches = dispatches.filter(
            Q(batch_number__icontains=search) |
            Q(items__scanned_order_id__icontains=search)
        ).distinct()
    
    # Date filters
    if date_from:
        dispatches = dispatches.filter(created_at__date__gte=date_from)
    
    if date_to:
        dispatches = dispatches.filter(created_at__date__lte=date_to)
    
    context = {
        'dispatches': dispatches,
        'logistics_choices': Dispatch.LOGISTICS_CHOICES,
        'status_choices': Dispatch.STATUS_CHOICES,
        'search': search,
        'logistics_filter': logistics_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'dispatch_list.html', context)


@login_required
@permission_required('can_view_dispatch')
def dispatch_detail(request, pk):
    """View single dispatch details"""
    dispatch = get_object_or_404(Dispatch.objects.prefetch_related('items__order'), pk=pk)
    
    context = {
        'dispatch': dispatch,
    }
    
    return render(request, 'dispatch_detail.html', context)


@login_required
@permission_required('can_delete_dispatch')
def dispatch_delete(request, pk):
    """Delete a dispatch"""
    dispatch = get_object_or_404(Dispatch, pk=pk)
    
    if request.method == 'POST':
        batch_number = dispatch.batch_number
        dispatch.delete()
        messages.success(request, f'Dispatch {batch_number} has been deleted.')
        return redirect('dispatch_list')
    
    return redirect('dispatch_detail', pk=pk)

@login_required
@permission_required('can_view_dispatch')
def dispatch_bulk_action(request):
    """Handle bulk actions on dispatches"""
    if request.method == 'POST':
        dispatch_ids = request.POST.getlist('dispatch_ids')
        action = request.POST.get('bulk_action')
        
        if not dispatch_ids:
            messages.error(request, 'No dispatches selected!')
            return redirect('dispatch_list')
        
        try:
            dispatches = Dispatch.objects.filter(id__in=dispatch_ids, is_deleted=False)
            count = dispatches.count()
            
            if count == 0:
                messages.error(request, 'No valid dispatches found!')
                return redirect('dispatch_list')
            
            if action == 'move_to_trash':
                dispatches.update(
                    is_deleted=True,
                    deleted_by=request.user,
                    deleted_at=timezone.now()
                )
                messages.success(request, f'✅ {count} dispatch(es) moved to trash!')
                
            elif action == 'change_status':
                new_status = request.POST.get('new_status')
                if not new_status:
                    messages.error(request, 'Please select a new status!')
                    return redirect('dispatch_list')
                
                dispatches.update(status=new_status)
                messages.success(request, f'✅ {count} dispatch(es) status updated to {new_status}!')
                
            elif action == 'export_excel':
                # Export functionality (you can implement this later)
                messages.info(request, 'Export functionality coming soon!')
                
            else:
                messages.error(request, 'Invalid action selected!')
                
        except Exception as e:
            messages.error(request, f'Error performing bulk action: {str(e)}')
    
    return redirect('dispatch_list')



# inventory dashboard
@login_required
@permission_required('can_view_inventory')
def inventory_dashboard(request):
    """Modern Inventory Dashboard with Analytics - COMPLETE VERSION"""
    try:
        from datetime import datetime, timedelta
        from django.db.models import Sum
        
        # Get user's products
        products = Product.objects.filter(user=request.user)
        
        # Stock Statistics
        total_products = products.count()
        in_stock = products.filter(stock__gt=10).count()
        low_stock = products.filter(stock__lte=10, stock__gt=0).count()
        out_of_stock = products.filter(stock=0).count()
        
        # Product Variations Stock
        variations = ProductVariation.objects.filter(product__user=request.user)
        total_variations = variations.count()
        variations_in_stock = variations.filter(stock__gt=10).count()
        variations_low_stock = variations.filter(stock__lte=10, stock__gt=0).count()
        variations_out_of_stock = variations.filter(stock=0).count()
        
        # Calculate Total Stock Value
        total_stock_value = sum(p.stock * p.price for p in products if p.stock > 0)
        
        # Total Stock Units
        total_stock_units = sum(p.stock for p in products)
        
        # Low Stock Products (for alerts table)
        low_stock_products = products.filter(stock__lte=10, stock__gt=0).order_by('stock')[:10]
        
        # Out of Stock Products
        out_of_stock_products = products.filter(stock=0).order_by('name')[:10]
        
        # Top Products by Stock Value
        products_with_value = [{'product': p, 'value': p.stock * p.price} for p in products if p.stock > 0]
        top_products = sorted(products_with_value, key=lambda x: x['value'], reverse=True)[:10]
        
        # Recent Dispatched Orders
        recent_dispatched_orders = Order.objects.filter(
            order_status='dispatched',
            created_by=request.user
        ).order_by('-dispatch_date')[:10]
        
        # ✅ Category-wise Stock Distribution WITH CHART DATA
        categories = Category.objects.all()
        category_stock = []
        category_labels = []
        category_data = []
        total_category_value = 0
        
        for cat in categories:
            cat_products = products.filter(category=cat)
            cat_stock = sum(p.stock for p in cat_products)
            cat_value = sum(p.stock * p.price for p in cat_products if p.stock > 0)
            
            if cat_stock > 0:
                category_stock.append({
                    'category': cat.name,
                    'stock': cat_stock,
                    'products': cat_products.count(),
                    'value': cat_value,
                    'percentage': 0  # Will calculate below
                })
                category_labels.append(cat.name)
                category_data.append(cat_stock)
                total_category_value += cat_value
        
        # Calculate percentages for progress bars
        for cat in category_stock:
            if total_category_value > 0:
                cat['percentage'] = (cat['value'] / total_category_value) * 100
            else:
                cat['percentage'] = 0
        
        # ✅ Stock Movement Data (Last 30 Days)
        today = timezone.now().date()
        movement_labels = []
        stock_in_data = []
        stock_out_data = []
        
        for i in range(29, -1, -1):
            date = today - timedelta(days=i)
            movement_labels.append(date.strftime('%b %d'))
            
            # Stock In for this date
            try:
                stock_ins_day = StockIn.objects.filter(
                    created_by=request.user,
                    created_at__date=date
                ).aggregate(total=Sum('total_quantity'))['total'] or 0
            except:
                stock_ins_day = 0
            
            # Stock Out (from dispatched orders) for this date
            try:
                orders_day = Order.objects.filter(
                    created_by=request.user,
                    order_status='dispatched',
                    dispatch_date__date=date
                )
                
                stock_out_day = 0
                for order in orders_day:
                    try:
                        stock_out_day += sum(item.quantity for item in order.items.all())
                    except:
                        pass
            except:
                stock_out_day = 0
            
            stock_in_data.append(stock_ins_day)
            stock_out_data.append(stock_out_day)
        
        # ✅ Stock Turnover Rate
        total_sold_30days = sum(stock_out_data)
        avg_inventory = total_stock_units if total_stock_units > 0 else 1
        stock_turnover_rate = (total_sold_30days / avg_inventory) if avg_inventory > 0 else 0
        
        # ✅ Dead Stock (No movement in 90 days)
        ninety_days_ago = today - timedelta(days=90)
        try:
            dead_stock_count = products.filter(
                updated_at__lt=ninety_days_ago,
                stock__gt=0
            ).count()
        except:
            dead_stock_count = 0
        
        # ✅ SAFE QUERY - Get Recent Stock In Transactions
        recent_stock_ins = []
        try:
            stock_ins_qs = StockIn.objects.filter(
                created_by=request.user
            ).only('id', 'reference_number', 'stock_in_type', 'supplier_name', 'created_at', 'total_quantity').order_by('-created_at')[:10]
            
            for stock_in in stock_ins_qs:
                try:
                    # Safely get total_cost
                    try:
                        total_cost = float(stock_in.total_cost) if stock_in.total_cost else 0.0
                    except:
                        total_cost = 0.0
                    
                    # Create safe dict
                    recent_stock_ins.append({
                        'id': stock_in.id,
                        'reference_number': stock_in.reference_number,
                        'stock_in_type': stock_in.stock_in_type,
                        'get_stock_in_type_display': stock_in.get_stock_in_type_display(),
                        'supplier_name': stock_in.supplier_name,
                        'total_quantity': stock_in.total_quantity,
                        'total_cost': total_cost,
                        'created_at': stock_in.created_at,
                        'items_count': stock_in.items.count(),
                    })
                except Exception as e:
                    print(f"⚠️ Skipping bad stock_in {stock_in.id}: {e}")
                    continue
                    
        except Exception as e:
            print(f"⚠️ Error loading stock ins: {e}")
            recent_stock_ins = []
        
        # Stock Status Distribution for Charts
        stock_chart_data = {
            'labels': ['In Stock', 'Low Stock', 'Out of Stock'],
            'data': [in_stock, low_stock, out_of_stock],
            'colors': ['#10b981', '#f59e0b', '#ef4444']
        }
        
        context = {
            # Basic Stats
            'total_products': total_products,
            'in_stock': in_stock,
            'low_stock': low_stock,
            'out_of_stock': out_of_stock,
            
            # Variations Stats
            'total_variations': total_variations,
            'variations_in_stock': variations_in_stock,
            'variations_low_stock': variations_low_stock,
            'variations_out_of_stock': variations_out_of_stock,
            
            # Value Stats
            'total_stock_value': total_stock_value,
            'total_stock_units': total_stock_units,
            'stock_turnover_rate': stock_turnover_rate,
            'dead_stock_count': dead_stock_count,
            
            # Product Lists
            'low_stock_products': low_stock_products,
            'out_of_stock_products': out_of_stock_products,
            'top_products': top_products,
            'recent_dispatched_orders': recent_dispatched_orders,
            
            # Category Data
            'category_stock': category_stock,
            
            # Stock In Transactions
            'recent_stock_ins': recent_stock_ins,
            
            # ✅ Chart Data (JSON encoded for JavaScript)
            'stock_chart_data': json.dumps(stock_chart_data),
            'category_labels': json.dumps(category_labels),
            'category_data': json.dumps(category_data),
            'movement_labels': json.dumps(movement_labels),
            'stock_in_data': json.dumps(stock_in_data),
            'stock_out_data': json.dumps(stock_out_data),
        }
        
        return render(request, 'inventory_dashboard.html', context)
        
    except Exception as e:
        print(f"❌ Error in inventory_dashboard: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Error loading inventory dashboard: {str(e)}')
        return redirect('dashboard')

# stock in 
@login_required
@permission_required('can_manage_inventory')
def stock_in_create(request):
    """Create new stock in transaction - BULLETPROOF VERSION"""
    
    if request.method == 'POST':
        try:
            # Get form data
            stock_in_type = request.POST.get('stock_in_type', 'purchase')
            supplier_name = request.POST.get('supplier_name', '').strip()
            notes = request.POST.get('notes', '').strip()
            items_json = request.POST.get('items', '[]')
            
            print(f"📦 Items JSON: {items_json}")
            
            # Parse items
            items = json.loads(items_json)
            
            if not items:
                messages.error(request, 'Please add at least one product')
                return redirect('stock_in_create')
            
            # Create StockIn WITHOUT totals first
            stock_in = StockIn.objects.create(
                stock_in_type=stock_in_type,
                supplier_name=supplier_name,
                notes=notes,
                created_by=request.user,
                total_quantity=0,
                total_cost=0
            )
            
            print(f"✅ Created StockIn: {stock_in.reference_number}")
            
            total_qty = 0
            total_cost = 0.0
            
            # Process items
            for idx, item in enumerate(items, 1):
                try:
                    product_id = int(item.get('product_id', 0))
                    quantity = int(item.get('quantity', 0))
                    
                    # ✅ Parse unit_cost as FLOAT first
                    unit_cost_str = str(item.get('unit_cost', '0')).strip()
                    try:
                        unit_cost_float = float(unit_cost_str)
                    except:
                        unit_cost_float = 0.0
                    
                    if quantity <= 0 or product_id <= 0:
                        continue
                    
                    # Get product
                    product = Product.objects.get(id=product_id, user=request.user)
                    
                    # Get variation if exists
                    variation = None
                    variation_id = item.get('variation_id')
                    if variation_id:
                        variation = ProductVariation.objects.get(id=int(variation_id))
                    
                    # Calculate total as FLOAT
                    item_total = quantity * unit_cost_float
                    
                    # Create item
                    StockInItem.objects.create(
                        stock_in=stock_in,
                        product=product,
                        product_variation=variation,
                        quantity=quantity,
                        unit_cost=unit_cost_float,
                        total_cost=item_total,
                        notes=item.get('notes', '')
                    )
                    
                    # Update stock
                    if variation:
                        variation.stock += quantity
                        if variation.stock > 0:
                            variation.status = 'active'
                        variation.save()
                    else:
                        product.stock += quantity
                        if product.stock > 0:
                            product.stock_status = 'in_stock'
                        product.save()
                    
                    total_qty += quantity
                    total_cost += item_total
                    
                    print(f"  ✅ Item {idx}: {product.name} x{quantity} @ {unit_cost_float}")
                    
                except Exception as e:
                    print(f"  ⚠️ Item {idx} error: {e}")
                    continue
            
            # Update totals
            stock_in.total_quantity = total_qty
            stock_in.total_cost = total_cost
            stock_in.save()
            
            messages.success(
                request,
                f'✅ Stock In {stock_in.reference_number} created! '
                f'{total_qty} items added worth Rs {total_cost:.2f}'
            )
            
            return redirect('inventory_dashboard')
            
        except json.JSONDecodeError:
            messages.error(request, 'Invalid data format')
        except Exception as e:
            print(f"❌ Error: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error: {str(e)}')
        
        return redirect('stock_in_create')
    
    # GET request
    products = Product.objects.filter(user=request.user, is_active=True).order_by('name')
    return render(request, 'stock_in_create.html', {'products': products})

@login_required
def stock_in_detail(request, stock_in_id):
    """View stock in transaction details - COMPLETE FIXED VERSION"""
    try:
        from django.db import connection
        from django.contrib.auth.models import User
        
        # ✅ Get Stock In basic data using RAW SQL
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    id, reference_number, stock_in_type, supplier_name, 
                    notes, total_quantity, created_at, created_by_id
                FROM dashboard_stockin
                WHERE id = %s AND created_by_id = %s
            """, [stock_in_id, request.user.id])
            
            row = cursor.fetchone()
            
            if not row:
                messages.error(request, 'Stock In transaction not found')
                return redirect('inventory_dashboard')
            
            # Create StockIn object from raw data
            class StockInData:
                def __init__(self, data):
                    self.id = data[0]
                    self.reference_number = data[1]
                    self.stock_in_type = data[2]
                    self.supplier_name = data[3]
                    self.notes = data[4]
                    self.total_quantity = data[5]
                    self.created_at = data[6]
                    self.created_by_id = data[7]
                
                def get_stock_in_type_display(self):
                    types = {
                        'purchase': 'Purchase Order',
                        'return': 'Customer Return',
                        'adjustment': 'Stock Adjustment',
                        'transfer': 'Transfer In',
                        'other': 'Other',
                    }
                    return types.get(self.stock_in_type, self.stock_in_type.title())
            
            stock_in = StockInData(row)
            
            # Get created_by user
            try:
                stock_in.created_by = User.objects.get(id=stock_in.created_by_id)
            except User.DoesNotExist:
                stock_in.created_by = type('obj', (object,), {'username': 'Unknown'})()
        
        # ✅ Get items using RAW SQL with product info
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    si.id, 
                    si.product_id, 
                    si.product_variation_id, 
                    si.quantity, 
                    si.notes,
                    p.name as product_name, 
                    p.slug as product_slug, 
                    p.image as product_image
                FROM dashboard_stockinitem si
                JOIN dashboard_product p ON si.product_id = p.id
                WHERE si.stock_in_id = %s
                ORDER BY si.id
            """, [stock_in_id])
            
            item_rows = cursor.fetchall()
        
        # Build items data list
        items_data = []
        total_cost_sum = 0.0
        
        for row in item_rows:
            item_id = row[0]
            product_id = row[1]
            variation_id = row[2]
            quantity = row[3]
            notes = row[4]
            product_name = row[5]
            product_slug = row[6]
            product_image = row[7]
            
            # ✅ Get SAFE costs using RAW SQL
            with connection.cursor() as cursor2:
                cursor2.execute("""
                    SELECT unit_cost, total_cost
                    FROM dashboard_stockinitem
                    WHERE id = %s
                """, [item_id])
                
                cost_row = cursor2.fetchone()
                
                if cost_row:
                    try:
                        safe_unit_cost = float(cost_row[0]) if cost_row[0] else 0.0
                    except (TypeError, ValueError, Decimal.InvalidOperation):
                        safe_unit_cost = 0.0
                    
                    try:
                        safe_total_cost = float(cost_row[1]) if cost_row[1] else 0.0
                    except (TypeError, ValueError, Decimal.InvalidOperation):
                        safe_total_cost = 0.0
                else:
                    safe_unit_cost = 0.0
                    safe_total_cost = 0.0
            
            # ✅ Get variation name if exists
            variation_name = None
            if variation_id:
                try:
                    variation = ProductVariation.objects.get(id=variation_id)
                    
                    # Get variation attributes
                    attribute_links = variation.attribute_values.select_related(
                        'attribute_value__attribute'
                    ).all()
                    
                    if attribute_links.exists():
                        variation_parts = []
                        for link in attribute_links:
                            attr = link.attribute_value.attribute.name
                            val = link.attribute_value.value
                            variation_parts.append(f"{attr}: {val}")
                        variation_name = " | ".join(variation_parts)
                    else:
                        variation_name = variation.sku
                except ProductVariation.DoesNotExist:
                    variation_name = f"Variation #{variation_id}"
                except Exception as e:
                    print(f"⚠️ Error getting variation {variation_id}: {e}")
                    variation_name = f"Variation #{variation_id}"
            
            # ✅ Create clean item object
            class ItemData:
                def __init__(self):
                    self.id = item_id
                    self.quantity = quantity
                    self.notes = notes if notes else ""
                    
                    # Create product sub-object
                    class ProductObj:
                        def __init__(self):
                            self.id = product_id
                            self.name = product_name
                            self.slug = product_slug
                            
                            # Handle image
                            if product_image:
                                class ImageObj:
                                    def __init__(self, url):
                                        self.url = url
                                self.image = ImageObj(product_image)
                            else:
                                self.image = None
                    
                    self.product = ProductObj()
            
            item = ItemData()
            
            # Add to total
            total_cost_sum += safe_total_cost
            
            # Add to items list
            items_data.append({
                'item': item,
                'variation_name': variation_name,
                'safe_unit_cost': safe_unit_cost,
                'safe_total_cost': safe_total_cost,
            })
        
        # ✅ Get safe total cost from StockIn table
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT total_cost
                FROM dashboard_stockin
                WHERE id = %s
            """, [stock_in_id])
            
            total_row = cursor.fetchone()
            
            if total_row and total_row[0]:
                try:
                    safe_total_cost = float(total_row[0])
                except (TypeError, ValueError, Decimal.InvalidOperation):
                    safe_total_cost = total_cost_sum
            else:
                safe_total_cost = total_cost_sum
        
        context = {
            'stock_in': stock_in,
            'items_data': items_data,
            'safe_total_cost': safe_total_cost,
        }
        
        return render(request, 'stock_in_detail.html', context)
        
    except Exception as e:
        print(f"❌ Error in stock_in_detail: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Error loading stock in details: {str(e)}')
        return redirect('inventory_dashboard')

@login_required
@require_http_methods(["GET"])
def api_get_product_for_stockin(request, product_id):
    """API to get product details including variations for stock in"""
    product = get_object_or_404(Product, id=product_id, user=request.user)
    
    data = {
        'id': product.id,
        'name': product.name,
        'type': product.product_type,
        'variations': []
    }
    
    if product.product_type == 'variable':
        variations = product.variations.filter(is_active=True).prefetch_related(
            'attribute_values__attribute_value__attribute'
        ).order_by('sku')
        
        for var in variations:
            attrs = []
            for link in var.attribute_values.all():
                av = link.attribute_value
                attrs.append({
                    'name': av.attribute.name,
                    'value': av.value
                })
            
            data['variations'].append({
                'id': var.id,
                'sku': var.sku,
                'stock': var.stock,
                'price': str(var.price),
                'attributes': attrs
            })
    else:
        data['stock'] = product.stock
        data['price'] = str(product.price)
    
    return JsonResponse(data)


# ==================== CITY MANAGEMENT VIEWS ====================

@login_required
@admin_only
def city_management(request):
    """City management page - Add, edit, delete cities with valley status"""
    cities = City.objects.all().order_by('name')
    
    if request.method == 'POST':
        # Add new city
        if 'add_city' in request.POST:
            name = request.POST.get('city_name', '').strip().title()
            valley_status = request.POST.get('valley_status', 'valley')
            
            if name:
                city, created = City.objects.get_or_create(
                    name=name,
                    defaults={
                        'valley_status': valley_status,
                        'is_active': True
                    }
                )
                
                if created:
                    messages.success(request, f'✅ City "{name}" added successfully!')
                else:
                    messages.info(request, f'ℹ️ City "{name}" already exists')
            else:
                messages.error(request, 'Please enter a city name')
                
        # Bulk actions
        elif 'bulk_action' in request.POST:
            city_ids = request.POST.getlist('city_ids')
            action = request.POST.get('bulk_action')
            
            if city_ids and action:
                cities_to_update = City.objects.filter(id__in=city_ids)
                
                if action == 'delete':
                    count = cities_to_update.count()
                    cities_to_update.delete()
                    messages.success(request, f'✅ {count} city(s) deleted successfully!')
                    
                elif action in ['valley', 'out_valley']:
                    cities_to_update.update(valley_status=action)
                    messages.success(request, f'✅ {cities_to_update.count()} city(s) updated to {action.replace("_", " ").title()}!')
                    
                elif action == 'activate':
                    cities_to_update.update(is_active=True)
                    messages.success(request, f'✅ {cities_to_update.count()} city(s) activated!')
                    
                elif action == 'deactivate':
                    cities_to_update.update(is_active=False)
                    messages.success(request, f'✅ {cities_to_update.count()} city(s) deactivated!')
    
    # Get statistics
    total_cities = cities.count()
    valley_cities = cities.filter(valley_status='valley').count()
    out_valley_cities = cities.filter(valley_status='out_valley').count()
    
    context = {
        'cities': cities,
        'total_cities': total_cities,
        'valley_cities': valley_cities,
        'out_valley_cities': out_valley_cities,
        'valley_status_choices': City.VALLEY_STATUS_CHOICES,
    }
    
    return render(request, 'city_management.html', context)

@login_required
@admin_only
def city_edit(request, city_id):
    """Edit a city"""
    city = get_object_or_404(City, id=city_id)
    
    if request.method == 'POST':
        name = request.POST.get('city_name', '').strip().title()
        valley_status = request.POST.get('valley_status', 'valley')
        is_active = 'is_active' in request.POST
        
        if name:
            # Check if name already exists (excluding current city)
            if City.objects.filter(name=name).exclude(id=city.id).exists():
                messages.error(request, f'City "{name}" already exists!')
            else:
                city.name = name
                city.valley_status = valley_status
                city.is_active = is_active
                city.save()
                messages.success(request, f'✅ City "{name}" updated successfully!')
                return redirect('city_management')
        else:
            messages.error(request, 'Please enter a city name')
    
    context = {
        'city': city,
        'valley_status_choices': City.VALLEY_STATUS_CHOICES,
    }
    
    return render(request, 'city_edit.html', context)


@login_required
@admin_only
def city_delete(request, city_id):
    """Delete a city"""
    city = get_object_or_404(City, id=city_id)
    
    if request.method == 'POST':
        city_name = city.name
        city.delete()
        messages.success(request, f'✅ City "{city_name}" deleted successfully!')
        return redirect('city_management')
    
    return render(request, 'city_delete.html', {'city': city})

@login_required
@admin_only
def city_quick_add(request):
    """Quick add cities via AJAX"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            city_name = data.get('city_name', '').strip().title()
            valley_status = data.get('valley_status', 'valley')
            
            if not city_name:
                return JsonResponse({'success': False, 'message': 'City name is required'})
            
            city, created = City.objects.get_or_create(
                name=city_name,
                defaults={
                    'valley_status': valley_status,
                    'is_active': True
                }
            )
            
            if created:
                return JsonResponse({
                    'success': True,
                    'message': f'City "{city_name}" added successfully!',
                    'city': {
                        'id': city.id,
                        'name': city.name,
                        'valley_status': city.valley_status,
                        'is_active': city.is_active,
                    }
                })
            else:
                # If city exists but valley status is different, update it
                if city.valley_status != valley_status:
                    city.valley_status = valley_status
                    city.save()
                    return JsonResponse({
                        'success': True,
                        'message': f'City "{city_name}" already exists. Updated valley status.',
                        'city': {
                            'id': city.id,
                            'name': city.name,
                            'valley_status': city.valley_status,
                            'is_active': city.is_active,
                        }
                    })
                return JsonResponse({
                    'success': True,
                    'message': f'City "{city_name}" already exists',
                    'city': {
                        'id': city.id,
                        'name': city.name,
                        'valley_status': city.valley_status,
                        'is_active': city.is_active,
                    }
                })
                
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid data format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})

@login_required
@admin_only
def city_bulk_add(request):
    """Bulk add multiple cities at once via AJAX"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            data = json.loads(request.body)
            city_name = data.get('city_name', '').strip().title()
            valley_status = data.get('valley_status', 'valley')
            is_bulk = data.get('is_bulk', False)
            
            if not city_name:
                return JsonResponse({'success': False, 'message': 'City name is required'})
            
            # Check if city already exists
            existing_city = City.objects.filter(name__iexact=city_name).first()
            
            if existing_city:
                # Check if valley status needs update
                if existing_city.valley_status != valley_status:
                    existing_city.valley_status = valley_status
                    existing_city.save()
                    return JsonResponse({
                        'success': True,
                        'message': f'Updated: {city_name}',
                        'status': 'updated',
                        'city_id': existing_city.id
                    })
                else:
                    return JsonResponse({
                        'success': True,
                        'message': f'Skipped: {city_name} (already exists)',
                        'status': 'skipped',
                        'city_id': existing_city.id
                    })
            
            # Create new city
            city = City.objects.create(
                name=city_name,
                valley_status=valley_status,
                is_active=True
            )
            
            if is_bulk:
                return JsonResponse({
                    'success': True,
                    'message': f'Added: {city_name}',
                    'status': 'added',
                    'city_id': city.id
                })
            else:
                return JsonResponse({
                    'success': True,
                    'message': f'City "{city_name}" added successfully as {city.get_valley_status_display()}',
                    'status': 'added',
                    'city_id': city.id
                })
                
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid data format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


@login_required
@require_http_methods(["GET"])
def api_get_cities(request):
    """API to get all cities for dropdown in order create"""
    try:
        cities = City.objects.filter(is_active=True).order_by('name')
        
        city_list = []
        for city in cities:
            city_list.append({
                'id': city.id,
                'name': city.name,
                'valley_status': city.valley_status,
                'display_status': city.get_valley_status_display(),
                'in_out': 'IN' if city.valley_status == 'valley' else 'OUT'
            })
        
        return JsonResponse({
            'success': True, 
            'cities': city_list,
            'count': len(city_list)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error loading cities: {str(e)}',
            'cities': []
        })
        
@login_required
@require_http_methods(["GET"])
def api_get_city_valley_status(request):
    """API to get valley status for a city (for IN/OUT field in order create)"""
    city_name = request.GET.get('city', '').strip().title()
    
    if not city_name:
        return JsonResponse({'success': False, 'message': 'City name required'})
    
    try:
        # Try exact match first, then case-insensitive match
        city = City.objects.filter(
            Q(name=city_name) | Q(name__iexact=city_name),
            is_active=True
        ).first()
        
        if city:
            return JsonResponse({
                'success': True,
                'city': city.name,
                'valley_status': city.valley_status,
                'display_status': city.get_valley_status_display(),
                'in_out': 'IN' if city.valley_status == 'valley' else 'OUT',
                'exists': True,
                'city_id': city.id
            })
        else:
            # Check if it's in default valley cities
            default_valley_cities = [
                'Kathmandu', 'Lalitpur', 'Bhaktapur', 'Kirtipur', 
                'Thimi', 'Tokha', 'Budhanilkantha', 'Gokarneshwor'
            ]
            
            if city_name in default_valley_cities:
                return JsonResponse({
                    'success': True,
                    'city': city_name,
                    'valley_status': 'valley',
                    'display_status': 'Valley (Inside)',
                    'in_out': 'IN',
                    'exists': False,
                    'is_default': True,
                    'message': f'City "{city_name}" is a default Valley city'
                })
            else:
                # Check common Out Valley cities
                common_out_valley_cities = [
                    'Pokhara', 'Biratnagar', 'Birgunj', 'Dharan', 'Hetauda', 
                    'Butwal', 'Nepalgunj', 'Dhankuta', 'Janakpur', 'Dhangadhi'
                ]
                
                if city_name in common_out_valley_cities:
                    return JsonResponse({
                        'success': True,
                        'city': city_name,
                        'valley_status': 'out_valley',
                        'display_status': 'Out Valley (Outside)',
                        'in_out': 'OUT',
                        'exists': False,
                        'is_default': True,
                        'message': f'City "{city_name}" is a common Out Valley city'
                    })
                else:
                    # Default to Out Valley for unknown cities
                    return JsonResponse({
                        'success': True,
                        'city': city_name,
                        'valley_status': 'out_valley',
                        'display_status': 'Out Valley (Outside)',
                        'in_out': 'OUT',
                        'exists': False,
                        'is_default': True,
                        'message': f'City "{city_name}" not found in database. Marked as Out Valley by default.'
                    })
                    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error checking city: {str(e)}'
        })
        
# ==================== VALLEY STATUS API ENDPOINT ====================

@login_required
@require_http_methods(["GET"])
def get_valley_status(request):
    """API endpoint to detect if a city is in valley or out valley - FIXED VERSION"""
    city_name = request.GET.get('city', '').strip()
    
    if not city_name:
        return JsonResponse({
            'success': False,
            'message': 'City name is required'
        })
    
    try:
        # Try to find the city in database (case-insensitive)
        city = City.objects.filter(
            Q(name__iexact=city_name) | Q(name__iexact=city_name.title())
        ).first()
        
        if city:
            return JsonResponse({
                'success': True,
                'valley_status': city.valley_status,
                'display_status': city.get_valley_status_display(),
                'exists': True,
                'city': city.name,
                'in_out': 'in' if city.valley_status == 'valley' else 'out'
            })
        else:
            # If city not found, check patterns
            city_lower = city_name.lower()
            
            # Valley city patterns
            valley_patterns = [
                'kathmandu', 'lalitpur', 'bhaktapur', 'kirtipur', 'patan',
                'thamel', 'ktm', 'kath', 'mandu', 'valley'
            ]
            
            # Out-valley city patterns
            out_valley_patterns = [
                'birgunj', 'pokhara', 'dharan', 'biratnagar', 'dumla',
                'butwal', 'bhairahawa', 'nepalgunj', 'dhangadhi',
                'hetauda', 'janakpur', 'dhankuta', 'outside', 'out'
            ]
            
            # Check patterns
            is_valley = any(pattern in city_lower for pattern in valley_patterns)
            is_out_valley = any(pattern in city_lower for pattern in out_valley_patterns)
            
            if is_valley and not is_out_valley:
                status = 'valley'
                display = 'Valley (Pattern Match)'
                in_out = 'in'
            elif is_out_valley and not is_valley:
                status = 'out_valley'
                display = 'Out Valley (Pattern Match)'
                in_out = 'out'
            else:
                # Default to OUT for unknown cities (safer assumption)
                status = 'out_valley'
                display = 'Out Valley (Default)'
                in_out = 'out'
            
            return JsonResponse({
                'success': True,
                'valley_status': status,
                'display_status': display,
                'exists': False,
                'city': city_name,
                'in_out': in_out,
                'message': f'City detected via pattern matching as {display}'
            })
            
    except Exception as e:
        print(f"Error in get_valley_status: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Error detecting valley status: {str(e)}'
        })
        
        
# ==================== RETURN MANAGEMENT VIEWS (WITH TRASH) ====================

@login_required
@permission_required('can_view_returns')
def returns_dashboard(request):
    """Return management dashboard with statistics"""
    
    # Get filter parameters
    date_filter = request.GET.get('date_range', 'all')
    status_filter = request.GET.get('status', '')
    
    # Base queryset - exclude deleted
    returns = ReturnRequest.objects.filter(is_deleted=False).select_related('order', 'customer', 'created_by').all()
    
    # Apply date filter
    today = timezone.now().date()
    if date_filter == 'today':
        returns = returns.filter(created_at__date=today)
    elif date_filter == 'yesterday':
        yesterday = today - timedelta(days=1)
        returns = returns.filter(created_at__date=yesterday)
    elif date_filter == 'last_7_days':
        start = today - timedelta(days=7)
        returns = returns.filter(created_at__date__gte=start)
    elif date_filter == 'last_30_days':
        start = today - timedelta(days=30)
        returns = returns.filter(created_at__date__gte=start)
    elif date_filter == 'this_month':
        returns = returns.filter(created_at__year=today.year, created_at__month=today.month)
    
    # Apply status filter
    if status_filter:
        returns = returns.filter(return_status=status_filter)
    
    # Statistics
    total_returns = returns.count()
    pending_returns = returns.filter(return_status='pending').count()
    approved_returns = returns.filter(return_status='approved').count()
    received_returns = returns.filter(return_status='received').count()
    inspecting_returns = returns.filter(return_status='inspecting').count()
    refunded_returns = returns.filter(return_status='refunded').count()
    rejected_returns = returns.filter(return_status='rejected').count()
    
    total_refund_amount = returns.filter(
        return_status='refunded'
    ).aggregate(total=Sum('refund_amount'))['total'] or Decimal('0.00')
    
    # Return reasons breakdown
    reason_stats = returns.values('return_reason').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Recent returns
    recent_returns = returns.order_by('-created_at')[:10]
    
    # Trash count
    trash_count = ReturnRequest.objects.filter(is_deleted=True).count()
    
    context = {
        'returns': recent_returns,
        'total_returns': total_returns,
        'pending_returns': pending_returns,
        'approved_returns': approved_returns,
        'received_returns': received_returns,
        'inspecting_returns': inspecting_returns,
        'refunded_returns': refunded_returns,
        'rejected_returns': rejected_returns,
        'total_refund_amount': total_refund_amount,
        'reason_stats': reason_stats,
        'date_filter': date_filter,
        'status_filter': status_filter,
        'trash_count': trash_count,
    }
    
    return render(request, 'returns/dashboard.html', context)


@login_required
@permission_required('can_view_returns')
def returns_list(request):
    """List all return requests with filters (excluding trash)"""
    
    returns = ReturnRequest.objects.filter(is_deleted=False).select_related(
        'order', 'customer', 'created_by', 'approved_by'
    ).prefetch_related('items').all().order_by('-created_at')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        returns = returns.filter(
            Q(rma_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(customer_phone__icontains=search_query) |
            Q(order__order_number__icontains=search_query)
        )
    
    # Filters
    status_filter = request.GET.get('status', '')
    if status_filter:
        returns = returns.filter(return_status=status_filter)
    
    reason_filter = request.GET.get('reason', '')
    if reason_filter:
        returns = returns.filter(return_reason=reason_filter)
    
    refund_type_filter = request.GET.get('refund_type', '')
    if refund_type_filter:
        returns = returns.filter(refund_type=refund_type_filter)
    
    # Date filter
    date_filter = request.GET.get('date_range', '')
    today = timezone.now().date()
    
    if date_filter == 'today':
        returns = returns.filter(created_at__date=today)
    elif date_filter == 'yesterday':
        yesterday = today - timedelta(days=1)
        returns = returns.filter(created_at__date=yesterday)
    elif date_filter == 'last_7_days':
        start = today - timedelta(days=7)
        returns = returns.filter(created_at__date__gte=start)
    elif date_filter == 'last_30_days':
        start = today - timedelta(days=30)
        returns = returns.filter(created_at__date__gte=start)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(returns, 25)
    page_number = request.GET.get('page')
    returns_page = paginator.get_page(page_number)
    
    context = {
        'returns': returns_page,
        'search_query': search_query,
        'status_filter': status_filter,
        'reason_filter': reason_filter,
        'refund_type_filter': refund_type_filter,
        'date_filter': date_filter,
        'status_choices': ReturnRequest.RETURN_STATUS_CHOICES,
        'reason_choices': ReturnRequest.RETURN_REASON_CHOICES,
        'refund_type_choices': ReturnRequest.REFUND_TYPE_CHOICES,
    }
    
    return render(request, 'returns/list.html', context)


@login_required
@permission_required('can_create_returns')
def return_create(request):
    """Create return request with barcode scanning support"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get order ID
                order_id = request.POST.get('order_id')
                if not order_id:
                    messages.error(request, 'Order not selected')
                    return redirect('return_create')
                
                order = get_object_or_404(Order, id=order_id, order_status='delivered')
                
                # Get return details
                return_reason = request.POST.get('return_reason')
                refund_type = request.POST.get('refund_type', 'full_refund')
                customer_notes = request.POST.get('customer_notes', '')
                
                # Parse return items
                return_items_json = request.POST.get('return_items', '[]')
                return_items_data = json.loads(return_items_json)
                
                if not return_items_data:
                    messages.error(request, 'No items selected for return')
                    return redirect('return_create')
                
                # Calculate total refund
                total_refund = Decimal('0.00')
                for item_data in return_items_data:
                    order_item = OrderItem.objects.get(id=item_data['order_item_id'])
                    qty = int(item_data['quantity'])
                    total_refund += order_item.price * qty
                
                # Create return request
                return_request = ReturnRequest.objects.create(
                    order=order,
                    customer=order.customer,
                    customer_name=order.customer_name,
                    customer_phone=order.customer_phone,
                    customer_email=order.customer_email,
                    return_reason=return_reason,
                    refund_type=refund_type,
                    customer_notes=customer_notes,
                    total_amount=order.total_amount,
                    refund_amount=total_refund,
                    created_by=request.user,
                )
                
                # Create return items
                for item_data in return_items_data:
                    order_item = OrderItem.objects.get(id=item_data['order_item_id'])
                    qty = int(item_data['quantity'])
                    
                    ReturnItem.objects.create(
                        return_request=return_request,
                        order_item=order_item,
                        product=order_item.product,
                        product_variation=order_item.product_variation,
                        product_name=order_item.product_name,
                        product_sku=order_item.product_sku,
                        quantity=order_item.quantity,
                        price=order_item.price,
                        total=order_item.total,
                        return_quantity=qty,
                        refund_amount=order_item.price * qty
                    )
                
                # Log activity
                ReturnActivityLog.objects.create(
                    return_request=return_request,
                    user=request.user,
                    action_type='created',
                    description=f'Return request {return_request.rma_number} created for order {order.order_number}'
                )
                
                messages.success(request, f'Return request {return_request.rma_number} created successfully!')
                return redirect('returns_list')
                
        except Exception as e:
            messages.error(request, f'Error creating return: {str(e)}')
            print(f"Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return redirect('return_create')
    
    # GET request
    # Get recent delivered orders for scanning
    recent_orders = Order.objects.filter(
        order_status='delivered',
        is_deleted=False
    ).select_related('customer', 'created_by').prefetch_related(
        'items__product',
        'items__product_variation'
    ).order_by('-delivered_at')[:50]
    
    # If order_id is provided in GET, load that order
    order = None
    order_items = []
    if request.GET.get('order_id'):
        try:
            order = Order.objects.select_related('customer').prefetch_related(
                'items__product',
                'items__product_variation'
            ).get(id=request.GET.get('order_id'), order_status='delivered')
            order_items = order.items.all()
        except Order.DoesNotExist:
            messages.error(request, 'Order not found or not delivered')
    
    context = {
        'recent_orders': recent_orders,
        'order': order,
        'order_items': order_items,
        'reason_choices': ReturnRequest.RETURN_REASON_CHOICES,
        'refund_type_choices': ReturnRequest.REFUND_TYPE_CHOICES,
    }
    
    return render(request, 'returns/create.html', context)

@login_required
def api_get_order_by_barcode(request):
    """AJAX endpoint to fetch order data by barcode/order number"""
    barcode = request.GET.get('barcode', '').strip()
    
    if not barcode:
        return JsonResponse({'success': False, 'error': 'No barcode provided'})
    
    try:
        # Build query - search by order number
        order = Order.objects.select_related('customer', 'created_by').prefetch_related(
            'items__product',
            'items__product_variation'
        ).filter(
            order_number__iexact=barcode,  # Case-insensitive match
            is_deleted=False
        ).first()
        
        # If not found, return helpful error
        if not order:
            # Check if order exists but is deleted
            deleted_order = Order.objects.filter(order_number__iexact=barcode, is_deleted=True).first()
            if deleted_order:
                return JsonResponse({
                    'success': False,
                    'error': f'Order "{barcode}" is in trash'
                })
            
            # Check if order exists but not delivered
            pending_order = Order.objects.filter(order_number__iexact=barcode, is_deleted=False).exclude(order_status='delivered').first()
            if pending_order:
                return JsonResponse({
                    'success': False,
                    'error': f'Order "{barcode}" is not delivered yet (Status: {pending_order.order_status})'
                })
            
            return JsonResponse({
                'success': False,
                'error': f'Order "{barcode}" not found'
            })
        
        # Check if delivered
        if order.order_status != 'delivered':
            return JsonResponse({
                'success': False,
                'error': f'Order "{barcode}" is not delivered (Status: {order.order_status}). Only delivered orders can be returned.'
            })
        
        # Prepare order items
        items = []
        for item in order.items.all():
            # Get SKU and barcode
            sku = item.product_sku or ''
            barcode_val = ''
            
            # Try to get barcode from variation or product
            if item.product_variation:
                if hasattr(item.product_variation, 'barcode'):
                    barcode_val = item.product_variation.barcode or ''
                if not barcode_val and hasattr(item.product, 'barcode'):
                    barcode_val = item.product.barcode or ''
                if not sku:
                    sku = item.product_variation.sku
            else:
                if hasattr(item.product, 'barcode'):
                    barcode_val = item.product.barcode or ''
            
            items.append({
                'id': item.id,
                'product_name': item.product_name,
                'product_sku': sku,
                'product_barcode': barcode_val,
                'price': str(item.price),
                'quantity': item.quantity,
                'product_variation': item.product_variation.sku if item.product_variation else None
            })
        
        return JsonResponse({
            'success': True,
            'order': {
                'id': order.id,
                'order_number': order.order_number,
                'customer_name': order.customer_name,
                'customer_phone': order.customer_phone,
                'customer_email': order.customer_email or '',
                'created_at': order.created_at.strftime('%b %d, %Y'),
                'total_amount': str(order.total_amount),
                'items': items
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        })


@login_required
@permission_required('can_view_returns')
def return_detail(request, return_id):
    """View return request details and update status"""
    
    return_request = get_object_or_404(
        ReturnRequest.objects.select_related(
            'order', 'customer', 'created_by', 'approved_by', 'quality_checked_by'
        ).prefetch_related('items', 'activity_logs'),
        id=return_id,
        is_deleted=False  # ✅ Only show non-deleted returns
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        try:
            if action == 'approve':
                return_request.return_status = 'approved'
                return_request.approved_by = request.user
                return_request.approved_at = timezone.now()
                return_request.save()
                
                ReturnActivityLog.objects.create(
                    return_request=return_request,
                    user=request.user,
                    action_type='approved',
                    description=f'Return approved by {request.user.username}'
                )
                
                messages.success(request, '✅ Return request approved!')
                
            elif action == 'reject':
                rejection_reason = request.POST.get('rejection_reason', '')
                return_request.return_status = 'rejected'
                return_request.rejection_reason = rejection_reason
                return_request.approved_by = request.user
                return_request.approved_at = timezone.now()
                return_request.save()
                
                ReturnActivityLog.objects.create(
                    return_request=return_request,
                    user=request.user,
                    action_type='rejected',
                    description=f'Return rejected: {rejection_reason}'
                )
                
                messages.warning(request, '⚠️ Return request rejected!')
                
            elif action == 'mark_received':
                return_request.return_status = 'received'
                return_request.save()
                
                ReturnActivityLog.objects.create(
                    return_request=return_request,
                    user=request.user,
                    action_type='received',
                    description='Returned items received at warehouse'
                )
                
                messages.success(request, '✅ Return marked as received!')
                
            elif action == 'quality_check':
                condition = request.POST.get('condition_received')
                quality_notes = request.POST.get('quality_check_notes', '')
                
                return_request.return_status = 'inspecting'
                return_request.condition_received = condition
                return_request.quality_check_notes = quality_notes
                return_request.quality_checked_by = request.user
                return_request.quality_checked_at = timezone.now()
                return_request.save()
                
                ReturnActivityLog.objects.create(
                    return_request=return_request,
                    user=request.user,
                    action_type='quality_checked',
                    description=f'Quality check completed. Condition: {condition}'
                )
                
                messages.success(request, '✅ Quality check completed!')
                
            elif action == 'process_refund':
                refund_amount = Decimal(request.POST.get('refund_amount', '0'))
                restocking_fee = Decimal(request.POST.get('restocking_fee', '0'))
                
                return_request.refund_amount = refund_amount
                return_request.restocking_fee = restocking_fee
                return_request.return_status = 'refunded'
                return_request.refunded_at = timezone.now()
                return_request.save()
                
                # Restock items
                for item in return_request.items.all():
                    if item.product_variation:
                        item.product_variation.stock += item.return_quantity
                        item.product_variation.save()
                    else:
                        item.product.stock += item.return_quantity
                        item.product.save()
                    
                    item.restocked = True
                    item.restocked_at = timezone.now()
                    item.restocked_by = request.user
                    item.save()
                
                ReturnActivityLog.objects.create(
                    return_request=return_request,
                    user=request.user,
                    action_type='refunded',
                    description=f'Refund processed: Rs. {refund_amount} (Restocking fee: Rs. {restocking_fee})'
                )
                
                messages.success(request, f'✅ Refund of Rs. {refund_amount} processed successfully!')
                
            elif action == 'update_notes':
                admin_notes = request.POST.get('admin_notes', '')
                return_request.admin_notes = admin_notes
                return_request.save()
                
                ReturnActivityLog.objects.create(
                    return_request=return_request,
                    user=request.user,
                    action_type='notes_updated',
                    description='Admin notes updated'
                )
                
                messages.success(request, '✅ Notes updated!')
            
            return redirect('return_detail', return_id=return_request.id)
            
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
            return redirect('return_detail', return_id=return_request.id)
    
    return_items = return_request.items.all()
    activity_logs = return_request.activity_logs.all()[:20]
    
    context = {
        'return_request': return_request,
        'return_items': return_items,
        'activity_logs': activity_logs,
        'condition_choices': ReturnRequest.CONDITION_CHOICES,
    }
    
    return render(request, 'returns/detail.html', context)


# ✅ TRASH MANAGEMENT VIEWS

@login_required
@permission_required('can_delete_returns')
def return_trash(request, return_id):
    """Move return to trash (soft delete)"""
    return_request = get_object_or_404(ReturnRequest, id=return_id, is_deleted=False)
    
    if request.method == 'POST':
        return_request.soft_delete(request.user)
        
        ReturnActivityLog.objects.create(
            return_request=return_request,
            user=request.user,
            action_type='trashed',
            description=f'Return {return_request.rma_number} moved to trash by {request.user.username}'
        )
        
        messages.success(request, f'✅ Return {return_request.rma_number} moved to trash!')
        return redirect('returns_list')
    
    context = {'return_request': return_request}
    return render(request, 'returns/trash_confirm.html', context)


@login_required
@permission_required('can_view_returns')
def returns_trash_list(request):
    """View all trashed returns"""
    
    trashed_returns = ReturnRequest.objects.filter(is_deleted=True).select_related(
        'order', 'customer', 'created_by', 'deleted_by'
    ).order_by('-deleted_at')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        trashed_returns = trashed_returns.filter(
            Q(rma_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(customer_phone__icontains=search_query)
        )
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(trashed_returns, 25)
    page_number = request.GET.get('page')
    returns_page = paginator.get_page(page_number)
    
    context = {
        'trashed_returns': returns_page,
        'search_query': search_query,
    }
    
    return render(request, 'returns/trash_list.html', context)


@login_required
@permission_required('can_delete_returns')
def return_restore(request, return_id):
    """Restore return from trash"""
    return_request = get_object_or_404(ReturnRequest, id=return_id, is_deleted=True)
    
    if request.method == 'POST':
        return_request.restore()
        
        ReturnActivityLog.objects.create(
            return_request=return_request,
            user=request.user,
            action_type='restored',
            description=f'Return {return_request.rma_number} restored from trash by {request.user.username}'
        )
        
        messages.success(request, f'✅ Return {return_request.rma_number} restored successfully!')
        return redirect('return_detail', return_id=return_request.id)
    
    context = {'return_request': return_request}
    return render(request, 'returns/restore_confirm.html', context)


@login_required
@admin_only
def return_permanent_delete(request, return_id):
    """Permanently delete return (Admin only)"""
    return_request = get_object_or_404(ReturnRequest, id=return_id, is_deleted=True)
    
    if request.method == 'POST':
        rma_number = return_request.rma_number
        return_request.delete()  # Permanent delete
        
        messages.success(request, f'✅ Return {rma_number} permanently deleted!')
        return redirect('returns_trash_list')
    
    context = {'return_request': return_request}
    return render(request, 'returns/permanent_delete_confirm.html', context)


@login_required
@admin_only
def returns_empty_trash(request):
    """Empty trash - permanently delete all trashed returns (Admin only)"""
    
    if request.method == 'POST':
        trashed_count = ReturnRequest.objects.filter(is_deleted=True).count()
        ReturnRequest.objects.filter(is_deleted=True).delete()
        
        messages.success(request, f'✅ {trashed_count} return(s) permanently deleted from trash!')
        return redirect('returns_trash_list')
    
    trashed_count = ReturnRequest.objects.filter(is_deleted=True).count()
    context = {'trashed_count': trashed_count}
    return render(request, 'returns/empty_trash_confirm.html', context)


# ✅ BULK ACTIONS

@login_required
@permission_required('can_delete_returns')
def returns_bulk_action(request):
    """Handle bulk actions on returns"""
    
    if request.method == 'POST':
        return_ids = request.POST.getlist('return_ids')
        action = request.POST.get('bulk_action')
        
        if not return_ids:
            messages.error(request, '❌ No returns selected!')
            return redirect('returns_list')
        
        returns = ReturnRequest.objects.filter(id__in=return_ids, is_deleted=False)
        count = returns.count()
        
        if action == 'trash':
            for return_request in returns:
                return_request.soft_delete(request.user)
                
                ReturnActivityLog.objects.create(
                    return_request=return_request,
                    user=request.user,
                    action_type='trashed',
                    description=f'Bulk moved to trash by {request.user.username}'
                )
            
            messages.success(request, f'✅ {count} return(s) moved to trash!')
            
        elif action == 'approve':
            returns.update(
                return_status='approved',
                approved_by=request.user,
                approved_at=timezone.now()
            )
            messages.success(request, f'✅ {count} return(s) approved!')
            
        elif action == 'reject':
            returns.update(
                return_status='rejected',
                approved_by=request.user,
                approved_at=timezone.now()
            )
            messages.success(request, f'✅ {count} return(s) rejected!')
        
        return redirect('returns_list')
    
    return redirect('returns_list')

# phone search API
@login_required
@require_http_methods(["GET"])
def search_customer_by_phone(request):
    """Search customer by phone number"""
    phone = request.GET.get('phone', '').strip()
    
    if not phone:
        return JsonResponse({'success': False, 'message': 'Phone number required'})
    
    try:
        # Search in Order model for customer with this phone
        from .models import Order
        
        # Get the most recent order with this phone number
        order = Order.objects.filter(customer_phone=phone).order_by('-created_at').first()
        
        if order:
            return JsonResponse({
                'success': True,
                'customer': {
                    'name': order.customer_name,
                    'email': order.customer_email or '',
                    'phone': order.customer_phone,
                    'address': order.shipping_address or '',
                    'landmark': order.landmark or '',
                    'branch_city': order.branch_city or '',
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'No customer found with this phone number'
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error searching customer: {str(e)}'
        })


# add custom product 

@login_required
@require_http_methods(["POST"])
def create_custom_product(request):
    """
    Create a custom product for quick sales.
    This product is saved to inventory and can be used immediately.
    """
    try:
        # Get form data
        name = request.POST.get('name', '').strip()
        price = request.POST.get('price', '0')
        stock = request.POST.get('stock', '1')
        sku = request.POST.get('sku', f'CUSTOM-{int(timezone.now().timestamp())}')
        category_id = request.POST.get('category', '')
        stock_status = request.POST.get('stock_status', 'in_stock')
        description = request.POST.get('description', '')
        
        # Validate required fields
        if not name:
            return JsonResponse({
                'success': False,
                'message': 'Product name is required'
            })
        
        # Convert and validate price and stock
        try:
            price = Decimal(price)
            stock = int(stock)
            
            if price <= 0:
                return JsonResponse({
                    'success': False,
                    'message': 'Price must be greater than 0'
                })
            
            if stock < 0:
                return JsonResponse({
                    'success': False,
                    'message': 'Stock cannot be negative'
                })
                
        except (ValueError, InvalidOperation):
            return JsonResponse({
                'success': False,
                'message': 'Invalid price or stock value'
            })
        
        # Get or create category
        category = None
        if category_id:
            try:
                category = Category.objects.get(id=category_id)
            except Category.DoesNotExist:
                pass
        
        # If no category provided or not found, get/create "Custom" category
        if not category:
            category, _ = Category.objects.get_or_create(
                name='Custom',
                defaults={
                    'slug': 'custom',
                }
            )
        
        # Generate unique slug
        base_slug = slugify(sku)
        slug = base_slug
        counter = 1
        while Product.objects.filter(slug=slug).exists():
            slug = f'{base_slug}-{counter}'
            counter += 1
        
        # Create product
        product = Product.objects.create(
            name=name,
            slug=slug,
            description=description or f'Custom product - {name}',
            price=price,
            cost_price=Decimal('0'),  # No cost for custom products
            stock=stock,
            category=category,
            stock_status=stock_status,
            product_type='simple',  # Always simple product
            barcode=sku,  # Use SKU as barcode
            is_active=True,
            is_deleted=False,
            user=request.user,
            is_custom_product=True,  # ✅ Mark as custom product
        )
        
        # Handle image upload
        if 'image' in request.FILES:
            product.image = request.FILES['image']
            product.save()
        
        # Return success response
        return JsonResponse({
            'success': True,
            'message': f'Custom product "{name}" created successfully',
            'product': {
                'id': product.id,
                'name': product.name,
                'price': str(product.price),
                'stock': product.stock,
                'sku': sku,
                'image': product.image.url if product.image else None
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Error creating custom product: {str(e)}'
        }, status=500)
        
    
    
    
    # my name is milan